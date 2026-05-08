"""
Builds two checklist deliverables (Excel + PDF) for the Tile Station platform:
  1. Website Features Checklist  (priority for launch verification)
  2. EPOS / Back-Office Features Checklist

Run:  python3 /app/checklists/build_checklists.py
Output: /app/checklists/*.xlsx, /app/checklists/*.pdf
        + mirror to /app/frontend/public/checklists/ for HTTPS download.

Each row carries:
  #  ·  Section  ·  Feature  ·  Priority (P0/P1/P2)  ·  Path  ·  Verified by E1  ·  User Sign-off
"""

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table as RLTable,
    TableStyle,
)


OUT_DIR = Path("/app/checklists")
PUBLIC_DIR = Path("/app/frontend/public/checklists")
OUT_DIR.mkdir(parents=True, exist_ok=True)
PUBLIC_DIR.mkdir(parents=True, exist_ok=True)
GENERATED_AT = datetime.now().strftime("%d %b %Y, %H:%M")


# =============================================================================
# WEBSITE FEATURES (priority — for launch cross-check)
# =============================================================================
WEBSITE_FEATURES = [
    # ---------------- Storefront — Homepage & Navigation ----------------
    ("Storefront — Homepage & Navigation", "Homepage hero, category grid, USP ribbon (Klarna · samples · collection · delivery)", "P0"),
    ("Storefront — Homepage & Navigation", "Editable Homepage Category tiles (image override, custom URL override, reset-to-default)", "P0"),
    ("Storefront — Homepage & Navigation", "Homepage carousel / video showroom section (admin toggle)", "P1"),
    ("Storefront — Homepage & Navigation", "Feature cards row (Free Delivery, Trade, Sample Service etc.)", "P1"),
    ("Storefront — Homepage & Navigation", "Header navigation menus (main yellow nav + footer quickLinks/customerService/legalLinks)", "P0"),
    ("Storefront — Homepage & Navigation", "Mega menu with 5 product groups: Tiles · Flooring · Materials · Tools & Accessories · Underfloor Heating", "P0"),
    ("Storefront — Homepage & Navigation", "Site Map (admin /admin/sitemap) with auto-link verification + Hyphen→Space migrator", "P1"),
    ("Storefront — Homepage & Navigation", "Maintenance advance-notice ribbon (≤24h before scheduled window)", "P1"),
    ("Storefront — Homepage & Navigation", "Whole-site maintenance switch (manual + scheduled window, auto enable/disable cron)", "P0"),

    # ---------------- Storefront — Product Browse & PDP ----------------
    ("Storefront — Product Browse & PDP", "Tile Collections page (/tiles) with filters (category, style, type, new, search)", "P0"),
    ("Storefront — Product Browse & PDP", "Dynamic page H1/subtitle by URL param (Marble/Wood/Stone/Wall/Floor/Outdoor/etc.)", "P0"),
    ("Storefront — Product Browse & PDP", "Collection Detail Page (/shop/collection/:slug) with variant + finish + size pills", "P0"),
    ("Storefront — Product Browse & PDP", "Dynamic proportional Size Pills (UI scales to actual tile dimensions)", "P0"),
    ("Storefront — Product Browse & PDP", "Volume Pricing tier table (e.g. 5%/10%/15% over thresholds)", "P0"),
    ("Storefront — Product Browse & PDP", "Trade Customer Box (login CTA) on PDP", "P0"),
    ("Storefront — Product Browse & PDP", "Add-to-cart with box-step quantity selector + sqm_per_box context", "P0"),
    ("Storefront — Product Browse & PDP", "Frequently-Bought-Together: installation essentials only (filtered, scheduled cache rebuild)", "P1"),
    ("Storefront — Product Browse & PDP", "“Show Actual Size” toggle + 1:1 print canvas (A4 PDF cheat-sheet route)", "P1"),
    ("Storefront — Product Browse & PDP", "Trade Bundle Booster CTA (free trade kit: tile + adhesive + grout samples)", "P1"),
    ("Storefront — Product Browse & PDP", "Free-sample CTA inside actual-size canvas (3-sample limit)", "P1"),
    ("Storefront — Product Browse & PDP", "Klarna OSM widget (“From £X/mo with Klarna”) — admin toggle", "P1"),
    ("Storefront — Product Browse & PDP", "SALE ribbon + bottom badge with compound discount math (consistent)", "P0"),
    ("Storefront — Product Browse & PDP", "Custom Product Labels (admin-managed colours)", "P1"),
    ("Storefront — Product Browse & PDP", "Wishlist (/shop/tile-wishlist) + Add-to-wishlist heart on PDP", "P1"),
    ("Storefront — Product Browse & PDP", "Compare Products tray, /shop/compare side-by-side spec table (admin toggle)", "P1"),
    ("Storefront — Product Browse & PDP", "Tile Calculator page", "P1"),
    ("Storefront — Product Browse & PDP", "Clearance / New Collections / Bathroom dedicated pages", "P1"),
    ("Storefront — Product Browse & PDP", "Stock status, trust badges, description formatting", "P1"),

    # ---------------- Storefront — Cart & Checkout ----------------
    ("Storefront — Cart & Checkout", "Tile Cart (/shop/tile-cart) with sticky free-delivery progress banner", "P0"),
    ("Storefront — Cart & Checkout", "Box-sized stepper (snaps quantity to whole boxes; m² and unit modes)", "P0"),
    ("Storefront — Cart & Checkout", "Tier upsell nudge (“Add 1 box to unlock X% off”)", "P1"),
    ("Storefront — Cart & Checkout", "Volume tier discount applied at cart/checkout (+ stacked Trade discount)", "P0"),
    ("Storefront — Cart & Checkout", "Per-line Tier Savings strip (volume + trade itemised)", "P1"),
    ("Storefront — Cart & Checkout", "Cart-level Total Savings hero strip (£ saved + % off retail)", "P1"),
    ("Storefront — Cart & Checkout", "Free Sample upsell (4 fulfillment modes: pack / separate / smart / hide-on-direct)", "P1"),
    ("Storefront — Cart & Checkout", "Cart Save Banner — guest email capture → WELCOME-* code (admin toggle)", "P1"),
    ("Storefront — Cart & Checkout", "Promo Code field (validates server-side, supports BACK/FRIEND/WELCOME/manual)", "P0"),
    ("Storefront — Cart & Checkout", "3-step Checkout (Cart → Address → Payment) with Tile Mountain receipt-style design", "P0"),
    ("Storefront — Cart & Checkout", "Billing Address (same-as-delivery checkbox + separate billing form)", "P0"),
    ("Storefront — Cart & Checkout", "Click & Collect option (delivery_method=collect, “Awaiting Collection” status)", "P0"),
    ("Storefront — Cart & Checkout", "Stripe Hosted Checkout — Card payments", "P0"),
    ("Storefront — Cart & Checkout", "Stripe — Klarna at checkout (≥£30) admin toggle", "P1"),
    ("Storefront — Cart & Checkout", "Klarna Express button on cart sidebar (one-tap)", "P1"),
    ("Storefront — Cart & Checkout", "PayPal Express button on cart sidebar (one-tap)", "P1"),
    ("Storefront — Cart & Checkout", "Apple Pay + Google Pay (Stripe ExpressCheckoutElement, .well-known/apple domain file)", "P1"),
    ("Storefront — Cart & Checkout", "Order success page + auto-tracking deep link", "P0"),
    ("Storefront — Cart & Checkout", "Tolerant guest-checkout schema (legacy carts with null fields accepted)", "P0"),
    ("Storefront — Cart & Checkout", "15-Minute Checkout Maintenance Warning (countdown replaces Pay CTA when window imminent)", "P1"),

    # ---------------- Storefront — Order Tracking ----------------
    ("Storefront — Order Tracking", "Public Track Order page (/shop/track) — order# + email lookup", "P0"),
    ("Storefront — Order Tracking", "Case-insensitive lookup + #-prefix tolerance + whitespace trim", "P0"),
    ("Storefront — Order Tracking", "Auto-track on URL params (?order=…&email=…) from email CTA", "P0"),
    ("Storefront — Order Tracking", "“Note from our team” banner (latest non-empty status_history.notes)", "P0"),
    ("Storefront — Order Tracking", "“Did you mean…?” recovery card (suggests recent paid orders for the email)", "P1"),

    # ---------------- Storefront — Sample Service ----------------
    ("Storefront — Sample Service", "Sample basket (3-sample limit) + TileSampleServicePage", "P1"),
    ("Storefront — Sample Service", "Sample checkout + sample order success", "P1"),
    ("Storefront — Sample Service", "Sample service CMS content (admin SampleServiceContent)", "P1"),

    # ---------------- Storefront — Customer / Trade Accounts ----------------
    ("Customer Accounts", "Customer Register (/shop/register) + Login (/shop/tile-login)", "P0"),
    ("Customer Accounts", "Customer Account dashboard (overview, orders, addresses)", "P0"),
    ("Customer Accounts", "Order History with Savings pill on each order", "P1"),
    ("Customer Accounts", "Forgot Password / Reset Password flow", "P0"),
    ("Trade Accounts", "Trade Login (/shop/trade/login) + Register (/shop/trade/register)", "P0"),
    ("Trade Accounts", "Trade Account dashboard (overview, orders, discount, addresses)", "P0"),
    ("Trade Accounts", "“Welcome Back, [Business Name]” banner with highlight chips", "P1"),
    ("Trade Accounts", "Trade Discount % + Tier badge + Credit Back £ balance", "P0"),
    ("Trade Accounts", "Order Again — one-click rebuy of last 3 paid orders", "P1"),
    ("Trade Accounts", "Activity Stream timeline (order/credit events) with deep links", "P1"),
    ("Trade Accounts", "Cart auto-reprice on Trade login/logout (toast)", "P0"),
    ("Trade Accounts", "Per-order savings pill in Trade Orders tab", "P1"),

    # ---------------- Marketing & Recovery Funnel ----------------
    ("Marketing & Recovery", "Welcome Popup (email capture → mints WELCOME-XXXXXX, emails code)", "P1"),
    ("Marketing & Recovery", "Refer-a-Friend (FRIEND-XXXXXX codes, /shop/refer page, share via WhatsApp/Email)", "P1"),
    ("Marketing & Recovery", "Abandoned Cart sequence — Day 0, Day 1, Last Chance emails", "P1"),
    ("Marketing & Recovery", "Abandoned Cart WhatsApp augmentation (Meta Cloud API, Day 1)", "P2"),
    ("Marketing & Recovery", "Promo Codes Admin (/admin/promo-codes) — mint/toggle/track redemptions", "P1"),
    ("Marketing & Recovery", "Marketing Funnel tile on /admin (captures, recovered £, codes minted/redeemed)", "P1"),
    ("Marketing & Recovery", "Weekly Digest email (Monday morning, configurable recipients/hour)", "P2"),
    ("Marketing & Recovery", "Storefront Messages admin (trade login/logout toasts copy + duration)", "P2"),

    # ---------------- Customer Communications ----------------
    ("Customer Communications", "Order confirmation email (branded, Resend, with savings pill)", "P0"),
    ("Customer Communications", "Order status update emails (paid, shipped, delivered, etc.) with deep-link CTA + admin notes", "P0"),
    ("Customer Communications", "Custom email from admin to customer (subject + body, with reply-to admin)", "P1"),
    ("Customer Communications", "Email History panel on order detail (capped 12, type pills, sender attribution)", "P1"),
    ("Customer Communications", "“Last contacted X ago” cue on Send-email button", "P2"),

    # ---------------- Static / Info Pages ----------------
    ("Static / Info Pages", "Privacy / Terms / FAQ / Returns / Delivery info pages", "P0"),
    ("Static / Info Pages", "Contact page", "P0"),
    ("Static / Info Pages", "Showrooms / Stores listing page", "P1"),
    ("Static / Info Pages", "Backwards-compat redirects (/shop/privacy, /shop/terms, /shop/faq, /shop/returns, /shop/delivery)", "P1"),

    # ---------------- Admin — Website / Storefront Management ----------------
    ("Admin — Website Management", "Website Hub landing (Quick Actions tiles)", "P1"),
    ("Admin — Website Management", "Homepage Manager (tiles, search filter, batch verify-all, copy share link, “Updated X ago”)", "P0"),
    ("Admin — Website Management", "Homepage Content Editor (hero, video showroom, copy)", "P1"),
    ("Admin — Website Management", "Navigation Menu Editor (header + footer, drag/drop, multi-level)", "P1"),
    ("Admin — Website Management", "Website Settings Editor (global settings, benefits bar, footer)", "P1"),
    ("Admin — Website Management", "Website Categories Manager", "P1"),
    ("Admin — Website Management", "Website Filters Manager + Specifications group scoping (bulk-assign)", "P1"),
    ("Admin — Website Management", "Collection Manager / Mappings / Page Settings / Detail Settings", "P1"),
    ("Admin — Website Management", "Info Pages Editor (Privacy/Terms/FAQ etc.)", "P1"),
    ("Admin — Website Management", "Storefront Features Admin (Compare, Welcome Popup, Cart Save Banner, Referrals — master toggles)", "P1"),
    ("Admin — Website Management", "Welcome Popup Admin (heading/copy/freq/email-coupon settings)", "P1"),
    ("Admin — Website Management", "Page Maintenance Admin (manual + scheduled window)", "P0"),
    ("Admin — Website Management", "Maintenance Tasks (DB Health snapshot + UI Health checks + migrators)", "P1"),
    ("Admin — Website Management", "Critical UI Health Panel (iframe smoke tests + Resend email alerts on failure)", "P1"),
    ("Admin — Website Management", "Site Map Manager + Hyphen→Space migrator", "P1"),
    ("Admin — Website Management", "Tools+Accessories DB consolidation migrator (idempotent)", "P1"),
    ("Admin — Website Management", "Bathroom Page Admin / Sample Service Admin", "P1"),

    # ---------------- Admin — Online Orders ----------------
    ("Admin — Online Orders", "Online Orders list (/admin/online-orders) with search + status filter + 25/page", "P0"),
    ("Admin — Online Orders", "“Today at a Glance” KPI strip (Orders/Revenue/Pending/Awaiting/Overdue)", "P0"),
    ("Admin — Online Orders", "Order detail modal (customer · payment · billing+delivery · items · history · status updater)", "P0"),
    ("Admin — Online Orders", "Status update modal with optional customer note + Quick Reply templates", "P0"),
    ("Admin — Online Orders", "Send Custom Email button + modal", "P1"),
    ("Admin — Online Orders", "Method column correctly shows Delivery vs Collect", "P0"),
    ("Admin — Online Orders", "Auto new-order audible notifier hook", "P1"),

    # ---------------- Admin — Marketing & Communication ----------------
    ("Admin — Marketing & Comms", "Communication Hub landing", "P1"),
    ("Admin — Marketing & Comms", "Abandoned Baskets Admin (KPIs, Settings, Baskets, Referrals tabs)", "P1"),
    ("Admin — Marketing & Comms", "Referrals tab (FRIEND codes, KPI cards, public share page link)", "P1"),
    ("Admin — Marketing & Comms", "Promo Codes Admin (mint, toggle, source filter, redeemed-£)", "P1"),
    ("Admin — Marketing & Comms", "Weekly Digest Admin (toggle, recipients, hour, live preview, send-now)", "P2"),
    ("Admin — Marketing & Comms", "WhatsApp Manager", "P2"),
    ("Admin — Marketing & Comms", "Email Inbox / Email Composer", "P2"),
    ("Admin — Marketing & Comms", "Notification Settings", "P2"),
    ("Admin — Marketing & Comms", "Live Chat Admin", "P2"),

    # ---------------- Admin — Live Visitors / Analytics ----------------
    ("Admin — Analytics", "Live Visitors page (heartbeat-based, by-page breakdown, recent activity)", "P1"),
    ("Admin — Analytics", "Website Analytics dashboard", "P1"),
    ("Admin — Analytics", "Website Sales Dashboard", "P1"),
    ("Admin — Analytics", "Marketing Funnel Card on main dashboard", "P1"),

    # ---------------- Admin — Permissions / RBAC ----------------
    ("Admin — RBAC", "Permissions Admin (/admin/permissions) — page + action grid", "P1"),
    ("Admin — RBAC", "Roles: super_admin / admin / manager / staff (system) + custom roles", "P1"),
    ("Admin — RBAC", "Column-level RBAC on Supplier Products (Cost/Live/Status/Actions)", "P1"),

    # ---------------- Storefront — Plumbing / Health ----------------
    ("Plumbing & Health", "Free-delivery threshold consolidated to single source (£499 admin-driven)", "P0"),
    ("Plumbing & Health", "savings_meta persisted on shop_orders (powers cart, email, dashboard pills)", "P1"),
    ("Plumbing & Health", "Trade reprice toasts (sessionStorage anchor — no duplicates on navigation)", "P0"),
    ("Plumbing & Health", "Auto-forward on Trade/Customer login pages when already authenticated", "P0"),
    ("Plumbing & Health", "data-testid coverage for critical UI elements (trade-customer-box, add-to-cart-btn, etc.)", "P0"),
    ("Plumbing & Health", "Box metadata backfill (sqm_per_box on supplier_products + tiles)", "P0"),
    ("Plumbing & Health", "Resend transactional email integration (verified sender online@tilestation.co.uk)", "P0"),
]


# =============================================================================
# EPOS / BACK-OFFICE FEATURES
# =============================================================================
EPOS_FEATURES = [
    # ---------------- Showroom EPOS — Tills / Sales ----------------
    ("Showroom EPOS — Tills", "Showroom EPOS terminal (/admin/showroom-epos) — barcode/SKU scan + manual entry", "P0"),
    ("Showroom EPOS — Tills", "Standard EPOS / Cash Counter terminal", "P0"),
    ("Showroom EPOS — Tills", "Cash Quotation + history", "P0"),
    ("Showroom EPOS — Tills", "Trade pricing applied automatically when trade customer linked", "P0"),
    ("Showroom EPOS — Tills", "Trade tier discount + Credit Back accrual", "P0"),
    ("Showroom EPOS — Tills", "Volume Pricing tier compatibility at till", "P0"),
    ("Showroom EPOS — Tills", "Multi-showroom support (4+ showrooms with stable string ids)", "P0"),
    ("Showroom EPOS — Tills", "Daily Sales Reconciliation card (Z-read: Cash & card vs Credit ledger)", "P0"),
    ("Showroom EPOS — Tills", "“Email me this” reconciliation (HTML + CSV attachment via Resend)", "P1"),
    ("Showroom EPOS — Tills", "Auto-scheduled daily reconciliation email (APScheduler hourly probe, configurable recipients/hour)", "P1"),
    ("Showroom EPOS — Tills", "Reconciliation audit log (last 5 dispatches with auto/manual badge)", "P1"),

    # ---------------- Invoicing & Documents ----------------
    ("Invoicing & Docs", "Invoice (create) + Invoice History", "P0"),
    ("Invoicing & Docs", "Showroom Invoice History (per-showroom view)", "P0"),
    ("Invoicing & Docs", "Proforma Invoice + history", "P0"),
    ("Invoicing & Docs", "Quotation + Quotation History", "P0"),
    ("Invoicing & Docs", "Refund + Refund History", "P0"),
    ("Invoicing & Docs", "Credit Note + Credit Note History", "P0"),
    ("Invoicing & Docs", "Document Storage", "P1"),
    ("Invoicing & Docs", "Soft-deleted invoices excluded from reconciliation pipeline", "P0"),

    # ---------------- Stock / Inventory ----------------
    ("Stock & Inventory", "Stock Hub landing", "P1"),
    ("Stock & Inventory", "Stock Allocation", "P0"),
    ("Stock & Inventory", "Stock Transfers (between showrooms)", "P0"),
    ("Stock & Inventory", "Stock Import", "P1"),
    ("Stock & Inventory", "Bulk Stock Edit", "P1"),
    ("Stock & Inventory", "Stocktake Report", "P1"),
    ("Stock & Inventory", "Stock Cost Report", "P1"),
    ("Stock & Inventory", "Reorder Suggestions", "P1"),
    ("Stock & Inventory", "To-Order Report", "P1"),
    ("Stock & Inventory", "Batch Tracking", "P1"),
    ("Stock & Inventory", "Delivery Check-In", "P1"),
    ("Stock & Inventory", "Delivery Management", "P1"),
    ("Stock & Inventory", "Always-In-Stock flag (per product, controllable from Supplier Products)", "P1"),

    # ---------------- Trade Customer Management ----------------
    ("Trade Management", "Trade Accounts list / detail (/admin/trade-accounts)", "P0"),
    ("Trade Management", "Trade tier ladder (Bronze/Silver/Gold etc. with discount %)", "P0"),
    ("Trade Management", "Customer Pricing overrides", "P1"),
    ("Trade Management", "Customer Invites flow", "P1"),
    ("Trade Management", "Trade List dashboard", "P1"),
    ("Trade Management", "Loyalty Dashboard (credit back, redemptions)", "P1"),

    # ---------------- Customers / CRM ----------------
    ("Customers / CRM", "Customers Hub landing", "P1"),
    ("Customers / CRM", "Quote Requests inbox", "P1"),
    ("Customers / CRM", "Bulk Inquiries", "P1"),
    ("Customers / CRM", "Tasks & Notes", "P1"),
    ("Customers / CRM", "Customer Account Settings (admin CMS)", "P1"),

    # ---------------- Products / Catalogue Admin ----------------
    ("Products / Catalogue", "Supplier Products monolith (~15k lines) — bulk editor", "P0"),
    ("Products / Catalogue", "Smart Select toolbar (m²/unit/none, sale/labels/new, tier default/custom/disabled)", "P1"),
    ("Products / Catalogue", "Floating Quick Actions bar (Apply Sale / Mark Not For Sale / Apply Description / Change Supplier / Archive)", "P1"),
    ("Products / Catalogue", "Bulk Category Editor (descriptions, scopes, categories, country of origin)", "P0"),
    ("Products / Catalogue", "Bulk Pricing (cost / list / sale / size-filtered)", "P0"),
    ("Products / Catalogue", "Bulk Pricing Unit (m² vs unit, type-filtered chips)", "P1"),
    ("Products / Catalogue", "Custom Tier Pricing per product (variable 2/3/4+ tiers)", "P1"),
    ("Products / Catalogue", "Sale & Labels modal (per-product checklist, type filter chips)", "P1"),
    ("Products / Catalogue", "Custom Product Labels (admin CRUD with colour picker)", "P1"),
    ("Products / Catalogue", "Price Lock / Name Lock (protects manual edits from sync overwrites)", "P0"),
    ("Products / Catalogue", "AI Description Generator (Display Name priority)", "P1"),
    ("Products / Catalogue", "Product Documents modal (PDFs)", "P1"),
    ("Products / Catalogue", "Quick Edit / Full Page Edit / Copy / Delete (cascades to tiles + products)", "P1"),
    ("Products / Catalogue", "Bulk Edit History + Presets", "P1"),
    ("Products / Catalogue", "Manage Options modal", "P1"),
    ("Products / Catalogue", "Spec/Filter group scoping (per product group)", "P1"),
    ("Products / Catalogue", "Specifications Manager + Filters Manager", "P1"),
    ("Products / Catalogue", "Categories Manager + Manage Categories", "P1"),
    ("Products / Catalogue", "Product Form (manual create)", "P1"),
    ("Products / Catalogue", "Clearance Products list", "P1"),
    ("Products / Catalogue", "New Collection Products list", "P1"),
    ("Products / Catalogue", "Trash + Restore", "P1"),
    ("Products / Catalogue", "Essentials Needing Photos tile (Sales Hub)", "P1"),

    # ---------------- Sync Hub & Suppliers ----------------
    ("Sync Hub & Suppliers", "Supplier Sync Dashboard / Sync Hub", "P0"),
    ("Sync Hub & Suppliers", "Supplier scrapers: Ceramica Impex / Verona / Wallcano / Splendour (resume + progress + timeout recovery)", "P0"),
    ("Sync Hub & Suppliers", "Force Stop / Reset / Validation Warnings UI", "P0"),
    ("Sync Hub & Suppliers", "Suppliers list (/admin/suppliers)", "P1"),
    ("Sync Hub & Suppliers", "Supplier Health Dashboard", "P1"),
    ("Sync Hub & Suppliers", "Supplier Images / Image Migration / Image Scraper", "P1"),
    ("Sync Hub & Suppliers", "Plus39 / Leporce supplier-specific image tools", "P2"),
    ("Sync Hub & Suppliers", "Wallcano Price Import", "P1"),
    ("Sync Hub & Suppliers", "Scraping Portal", "P1"),
    ("Sync Hub & Suppliers", "Null SKU migration (supplier_code → sku, startup)", "P0"),
    ("Sync Hub & Suppliers", "Cleanup Orphaned Tiles endpoint", "P1"),

    # ---------------- Reports / Sales Hub ----------------
    ("Reports / Sales Hub", "Sales Hub landing (Daily Reconciliation tile + KPI cards)", "P0"),
    ("Reports / Sales Hub", "Reports Hub landing", "P1"),
    ("Reports / Sales Hub", "Reports page (revenue, orders, products)", "P1"),
    ("Reports / Sales Hub", "Weekly Compare modal (Sun-Sat windows + quick-picks: Prior Week / YoY / 4 Weeks)", "P1"),
    ("Reports / Sales Hub", "Monthly Compare modal (parity with Weekly)", "P1"),
    ("Reports / Sales Hub", "Historical Sales — manual entries", "P1"),
    ("Reports / Sales Hub", "Showroom Dashboard (per-showroom)", "P1"),
    ("Reports / Sales Hub", "Staff Performance Dashboard", "P1"),
    ("Reports / Sales Hub", "Audit Trail page", "P1"),

    # ---------------- Order Management (legacy) ----------------
    ("Order Management", "Legacy Orders page (db.orders collection)", "P1"),
    ("Order Management", "Order Management workflow", "P1"),

    # ---------------- Staff & Auth ----------------
    ("Staff & Auth", "Admin Auth Page (/admin/auth)", "P0"),
    ("Staff & Auth", "Staff Register + Staff Invites", "P1"),
    ("Staff & Auth", "Staff PINs (till access)", "P1"),
    ("Staff & Auth", "User Management", "P1"),
    ("Staff & Auth", "Permissions Admin / Roles (RBAC)", "P1"),
    ("Staff & Auth", "Security Settings", "P1"),
    ("Staff & Auth", "Staff Chat", "P2"),

    # ---------------- Showrooms & Settings ----------------
    ("Showrooms & Settings", "Showrooms Manager (CRUD)", "P0"),
    ("Showrooms & Settings", "Settings Hub landing", "P1"),
    ("Showrooms & Settings", "Settings Page (general)", "P1"),
    ("Showrooms & Settings", "Pricing Settings", "P1"),
    ("Showrooms & Settings", "Tile Calculator Settings", "P1"),
    ("Showrooms & Settings", "Tiles Info CMS", "P1"),
    ("Showrooms & Settings", "Trade Account Settings (CMS)", "P1"),
    ("Showrooms & Settings", "Checkout Settings (delivery, payments incl. Klarna/PayPal/Wallet/Card)", "P0"),
    ("Showrooms & Settings", "Collection Detail Settings (delivery estimate, free threshold)", "P0"),
    ("Showrooms & Settings", "Collections Page Settings", "P1"),
    ("Showrooms & Settings", "Price Tickets / Print", "P1"),

    # ---------------- Health, Migrators, Plumbing ----------------
    ("Health & Migrators", "DB Health Snapshot (products, paid orders today, draft baskets, last paid order)", "P1"),
    ("Health & Migrators", "Critical UI Health checks (registry-driven, email alerts via Resend)", "P1"),
    ("Health & Migrators", "Hyphen→Space URL migrator (Site Map)", "P1"),
    ("Health & Migrators", "Tools + Accessories merge migrator", "P1"),
    ("Health & Migrators", "Legacy /shop/tiles path rewrite migrator", "P1"),
    ("Health & Migrators", "Showroom id backfill migrator", "P1"),
    ("Health & Migrators", "Box metadata backfill (sqm_per_box / tiles_per_box) script", "P0"),
    ("Health & Migrators", "Co-purchase cache nightly cron (FBT essentials)", "P1"),
    ("Health & Migrators", "Site maintenance window cron (auto enable/disable)", "P0"),
    ("Health & Migrators", "Abandoned cart reminders cron (every 15 min)", "P1"),
    ("Health & Migrators", "Weekly Digest cron (Mon 09:00 UTC default)", "P2"),
]


# =============================================================================
# URL/PATH RESOLVER (rule-based — first match wins)
# =============================================================================
URL_RULES: list[tuple[re.Pattern, str]] = [
    (re.compile(r"track order|track\b|tracking", re.I), "/shop/track"),
    (re.compile(r"refer-?a-?friend|/shop/refer|FRIEND-", re.I), "/shop/refer"),
    (re.compile(r"compare\b", re.I), "/shop/compare"),
    (re.compile(r"wishlist", re.I), "/shop/tile-wishlist"),
    (re.compile(r"sample (basket|service|cart|checkout)|samplecart|tilesamples", re.I), "/shop/sample-service"),
    (re.compile(r"order success|order-success|auto-tracking deep link", re.I), "/shop/order-success"),
    (re.compile(r"checkout\b", re.I), "/shop/tile-checkout"),
    (re.compile(r"\bcart\b", re.I), "/shop/tile-cart"),
    (re.compile(r"collection detail|/shop/collection|frequently-bought|size pills|trade customer box|actual size|trade bundle", re.I), "/shop/collection/Ridgeway%20Polished"),
    (re.compile(r"tile collections|/tiles\b|dynamic page h1", re.I), "/tiles"),
    (re.compile(r"clearance\b", re.I), "/shop/clearance"),
    (re.compile(r"new collection|new collections", re.I), "/shop/new-collections"),
    (re.compile(r"bathroom page|bathroom\b", re.I), "/shop/bathroom"),
    (re.compile(r"trade login|/shop/trade/login", re.I), "/shop/trade/login"),
    (re.compile(r"trade register|/shop/trade/register|trade signup", re.I), "/shop/trade/register"),
    (re.compile(r"trade account|trade dashboard|welcome back|order again|activity stream|reprice|trade discount", re.I), "/shop/trade/account"),
    (re.compile(r"customer register|/shop/register", re.I), "/shop/register"),
    (re.compile(r"customer login|/shop/tile-login", re.I), "/shop/tile-login"),
    (re.compile(r"customer account|order history|forgot password|reset password", re.I), "/shop/account"),
    (re.compile(r"contact page", re.I), "/shop/contact"),
    (re.compile(r"showroom epos|cash counter|standard epos|epos terminal|barcode/sku", re.I), "/admin/showroom-epos"),
    (re.compile(r"showrooms manager|showrooms? \(crud\)", re.I), "/admin/showrooms-manager"),
    (re.compile(r"showrooms?|stores", re.I), "/shop/stores"),
    (re.compile(r"info page|privacy|terms|faq|returns?|delivery info|backwards-compat redirect", re.I), "/shop/info/privacy"),
    (re.compile(r"calculator", re.I), "/shop/tile-calculator"),
    (re.compile(r"homepage|category grid|usp ribbon|hero|nav|mega menu|maintenance advance-notice|whole-site maintenance|carousel|feature cards", re.I), "/shop"),

    # ---------- Admin ----------
    (re.compile(r"online orders|today at a glance|order detail modal|status update modal|send custom email|method column|new-order audible", re.I), "/admin/online-orders"),
    (re.compile(r"homepage manager", re.I), "/admin/homepage-manager"),
    (re.compile(r"homepage content editor", re.I), "/admin/homepage-content"),
    (re.compile(r"navigation menu|navigation structure", re.I), "/admin/navigation-menu"),
    (re.compile(r"website settings editor", re.I), "/admin/website-settings"),
    (re.compile(r"website categories", re.I), "/admin/website-categories"),
    (re.compile(r"website filters|specifications group", re.I), "/admin/filters"),
    (re.compile(r"info pages editor", re.I), "/admin/info-pages"),
    (re.compile(r"storefront features", re.I), "/admin/storefront-features"),
    (re.compile(r"welcome popup admin", re.I), "/admin/welcome-popup"),
    (re.compile(r"page maintenance", re.I), "/admin/page-maintenance"),
    (re.compile(r"maintenance tasks|db health snapshot|ui health|migrator", re.I), "/admin/maintenance"),
    (re.compile(r"site map|hyphen", re.I), "/admin/sitemap"),
    (re.compile(r"bathroom page admin", re.I), "/admin/bathroom-page"),
    (re.compile(r"sample service.*(admin|cms)", re.I), "/admin/sample-service-content"),
    (re.compile(r"collection.*(manager|mappings|page settings|detail settings)", re.I), "/admin/collections-hub"),
    (re.compile(r"website hub", re.I), "/admin/website-hub"),
    (re.compile(r"communication hub", re.I), "/admin/communication-hub"),
    (re.compile(r"abandoned baskets|abandoned cart", re.I), "/admin/abandoned-baskets"),
    (re.compile(r"promo codes", re.I), "/admin/promo-codes"),
    (re.compile(r"weekly digest", re.I), "/admin/weekly-digest"),
    (re.compile(r"whatsapp", re.I), "/admin/whatsapp"),
    (re.compile(r"email inbox", re.I), "/admin/email-inbox"),
    (re.compile(r"email composer", re.I), "/admin/email-composer"),
    (re.compile(r"notification settings", re.I), "/admin/notifications"),
    (re.compile(r"live chat", re.I), "/admin/live-chat"),
    (re.compile(r"live visitors", re.I), "/admin/live-visitors"),
    (re.compile(r"website analytics", re.I), "/admin/website-analytics"),
    (re.compile(r"website sales dashboard", re.I), "/admin/website-sales"),
    (re.compile(r"marketing funnel|main dashboard", re.I), "/admin"),
    (re.compile(r"permissions", re.I), "/admin/permissions"),
    (re.compile(r"supplier products|smart select|quick actions bar|bulk category|bulk pricing|sale & labels|tier pricing|product labels|name lock|description generator|product documents|quick edit|bulk edit|manage options", re.I), "/admin/supplier-products"),
    (re.compile(r"specifications manager", re.I), "/admin/specifications"),
    (re.compile(r"categor(y|ies) manager|manage categories", re.I), "/admin/categories"),
    (re.compile(r"product form", re.I), "/admin/products/new"),
    (re.compile(r"clearance products", re.I), "/admin/clearance-products"),
    (re.compile(r"new collection products", re.I), "/admin/new-collection-products"),
    (re.compile(r"trash", re.I), "/admin/trash"),
    (re.compile(r"essentials needing photos|sales hub", re.I), "/admin/sales-hub"),
    (re.compile(r"showroom epos|cash counter", re.I), "/admin/showroom-epos"),
    (re.compile(r"epos terminal|standard epos", re.I), "/admin/epos"),
    (re.compile(r"cash quotation", re.I), "/admin/cash-quotation"),
    (re.compile(r"daily.*reconciliation|email me this|reconciliation audit|reconciliation email", re.I), "/admin/sales-hub"),
    (re.compile(r"invoice history", re.I), "/admin/invoice-history"),
    (re.compile(r"showroom invoice", re.I), "/admin/showroom-invoice-history"),
    (re.compile(r"invoice\b", re.I), "/admin/invoice"),
    (re.compile(r"proforma", re.I), "/admin/proforma-invoice"),
    (re.compile(r"quotation history", re.I), "/admin/quotation-history"),
    (re.compile(r"quotation", re.I), "/admin/quotation"),
    (re.compile(r"refund history", re.I), "/admin/refund-history"),
    (re.compile(r"refund", re.I), "/admin/refund"),
    (re.compile(r"credit note history", re.I), "/admin/credit-note-history"),
    (re.compile(r"credit note", re.I), "/admin/credit-note"),
    (re.compile(r"document storage", re.I), "/admin/document-storage"),
    (re.compile(r"stock hub", re.I), "/admin/stock-hub"),
    (re.compile(r"stock allocation", re.I), "/admin/stock-allocation"),
    (re.compile(r"stock transfers", re.I), "/admin/stock-transfers"),
    (re.compile(r"stock import", re.I), "/admin/stock-import"),
    (re.compile(r"bulk stock", re.I), "/admin/bulk-stock-edit"),
    (re.compile(r"stocktake", re.I), "/admin/stocktake-report"),
    (re.compile(r"stock cost", re.I), "/admin/stock-cost-report"),
    (re.compile(r"reorder suggestions", re.I), "/admin/reorder-suggestions"),
    (re.compile(r"to-?order report", re.I), "/admin/to-order-report"),
    (re.compile(r"batch tracking", re.I), "/admin/batch-tracking"),
    (re.compile(r"delivery check", re.I), "/admin/delivery-check-in"),
    (re.compile(r"delivery management", re.I), "/admin/delivery-management"),
    (re.compile(r"trade accounts list", re.I), "/admin/trade-accounts"),
    (re.compile(r"trade tier ladder", re.I), "/admin/trade-accounts"),
    (re.compile(r"customer pricing", re.I), "/admin/customer-pricing"),
    (re.compile(r"customer invites", re.I), "/admin/customer-invites"),
    (re.compile(r"trade list", re.I), "/admin/trade-list"),
    (re.compile(r"loyalty", re.I), "/admin/loyalty-dashboard"),
    (re.compile(r"customers hub", re.I), "/admin/customers-hub"),
    (re.compile(r"quote requests", re.I), "/admin/quote-requests"),
    (re.compile(r"bulk inquiries", re.I), "/admin/bulk-inquiries"),
    (re.compile(r"tasks.*notes", re.I), "/admin/tasks-notes"),
    (re.compile(r"customer account settings", re.I), "/admin/customer-account-settings"),
    (re.compile(r"sync hub|supplier sync", re.I), "/admin/sync-hub"),
    (re.compile(r"supplier scrapers|force stop|validation warnings|null sku|cleanup orphaned", re.I), "/admin/supplier-sync-dashboard"),
    (re.compile(r"\bsuppliers? list\b|^suppliers", re.I), "/admin/suppliers"),
    (re.compile(r"supplier health", re.I), "/admin/supplier-health"),
    (re.compile(r"supplier images|image migration|image scraper|leporce|plus39", re.I), "/admin/supplier-images"),
    (re.compile(r"wallcano price", re.I), "/admin/wallcano-price-import"),
    (re.compile(r"scraping portal", re.I), "/admin/scraping-portal"),
    (re.compile(r"reports hub", re.I), "/admin/reports-hub"),
    (re.compile(r"reports page|reports.*revenue|weekly compare|monthly compare", re.I), "/admin/reports"),
    (re.compile(r"historical sales", re.I), "/admin/historical-sales"),
    (re.compile(r"showroom dashboard", re.I), "/admin/showroom-dashboard"),
    (re.compile(r"staff performance", re.I), "/admin/staff-performance-dashboard"),
    (re.compile(r"audit trail", re.I), "/admin/audit-trail"),
    (re.compile(r"legacy orders|order management workflow", re.I), "/admin/orders"),
    (re.compile(r"admin auth", re.I), "/admin/auth"),
    (re.compile(r"staff register|staff invites", re.I), "/admin/staff-invites"),
    (re.compile(r"staff pins", re.I), "/admin/staff-pins"),
    (re.compile(r"user management", re.I), "/admin/user-management"),
    (re.compile(r"security settings", re.I), "/admin/security-settings"),
    (re.compile(r"staff chat", re.I), "/admin/staff-chat"),
    (re.compile(r"showrooms manager|showrooms? \(crud\)", re.I), "/admin/showrooms-manager"),
    (re.compile(r"settings hub", re.I), "/admin/settings-hub"),
    (re.compile(r"pricing settings", re.I), "/admin/pricing-settings"),
    (re.compile(r"tile calculator settings", re.I), "/admin/tile-calculator-settings"),
    (re.compile(r"tiles info", re.I), "/admin/tiles-info"),
    (re.compile(r"trade account settings", re.I), "/admin/trade-account-settings"),
    (re.compile(r"checkout settings", re.I), "/admin/checkout-settings"),
    (re.compile(r"collection detail settings", re.I), "/admin/collection-detail-settings"),
    (re.compile(r"collections page settings", re.I), "/admin/collections-page-settings"),
    (re.compile(r"price tickets", re.I), "/admin/price-tickets"),
]


def resolve_path(feature_text: str) -> str:
    for pattern, path in URL_RULES:
        if pattern.search(feature_text):
            return path
    return "—"


# =============================================================================
# ADMIN MANAGE-AT RESOLVER — where in the admin panel can you manage / edit
# this feature? For pure-admin rows this mirrors the storefront path; for
# storefront/customer-facing rows it points at the CMS / settings page that
# controls it.  First match wins.
# =============================================================================
ADMIN_RULES: list[tuple[re.Pattern, str]] = [
    # ----- Homepage / storefront chrome -----
    (re.compile(r"editable homepage category|homepage category tiles|homepage manager|category banner", re.I), "/admin/homepage-manager"),
    (re.compile(r"homepage carousel|video showroom|hero (slide|content)|homepage content", re.I), "/admin/homepage-content"),
    (re.compile(r"homepage hero|category grid|usp ribbon|feature cards", re.I), "/admin/homepage-content"),
    (re.compile(r"header navigation|footer .*links|mega menu|nav menu", re.I), "/admin/navigation-menu"),
    (re.compile(r"benefits bar|footer\b|website setting", re.I), "/admin/website-settings"),
    (re.compile(r"site map", re.I), "/admin/sitemap"),
    (re.compile(r"maintenance advance|whole-site maintenance|page maintenance|maintenance ribbon|maintenance window cron", re.I), "/admin/page-maintenance"),
    (re.compile(r"15-minute checkout maintenance|maintenance warning", re.I), "/admin/page-maintenance"),

    # ----- PDP (collection detail) — the page itself is CMS-driven via several admin tools -----
    (re.compile(r"size pills|trade customer box|frequently-bought-together|trade bundle booster|free-sample cta|klarna osm widget", re.I), "/admin/storefront-features"),
    (re.compile(r"volume pricing|tier pricing|tier table|tier discount", re.I), "/admin/pricing-settings"),
    (re.compile(r"sale ribbon|sale & labels|custom product labels|sale\b.*badge", re.I), "/admin/supplier-products"),
    (re.compile(r"add-to-cart with box|box-step|box metadata|sqm_per_box", re.I), "/admin/supplier-products"),
    (re.compile(r"show actual size|actual size|a4 pdf cheat", re.I), "/admin/collection-detail-settings"),
    (re.compile(r"trust badge|stock status|description formatting|tile detail|product browse|/shop/collection|collection detail", re.I), "/admin/collection-detail-settings"),
    (re.compile(r"tile collections page|/tiles\b|dynamic page h1|browse.*page", re.I), "/admin/collections-page-settings"),
    (re.compile(r"clearance products?|new collection products?|bathroom page|sample basket|sample checkout|sample order", re.I), "/admin/sample-service-content"),
    (re.compile(r"wishlist", re.I), "/admin/storefront-features"),
    (re.compile(r"compare\b", re.I), "/admin/storefront-features"),
    (re.compile(r"tile calculator settings", re.I), "/admin/tile-calculator-settings"),
    (re.compile(r"tile calculator(?! settings)", re.I), "/admin/tile-calculator-settings"),

    # ----- Cart & checkout -----
    (re.compile(r"\btile cart\b|/shop/tile-cart|sticky free-delivery progress", re.I), "/admin/checkout-settings"),
    (re.compile(r"box-sized stepper|box-step quantity|stepper", re.I), "/admin/supplier-products"),
    (re.compile(r"free-delivery threshold|delivery estimate", re.I), "/admin/collection-detail-settings"),
    (re.compile(r"tier upsell|per-line tier|cart-level total savings|free sample upsell", re.I), "/admin/pricing-settings"),
    (re.compile(r"cart save banner|welcome popup", re.I), "/admin/welcome-popup"),
    (re.compile(r"promo code field|promo code", re.I), "/admin/promo-codes"),
    (re.compile(r"3-step checkout|billing address|click & collect|stripe (hosted|—)|klarna (express|at checkout)|paypal express|apple pay|google pay|wallet|/admin/checkout-settings|order success|tolerant guest", re.I), "/admin/checkout-settings"),
    (re.compile(r"clearance|new collections|bathroom dedicated|new collection / bathroom", re.I), "/admin/website-categories"),
    (re.compile(r"referrals tab|friend codes|referral", re.I), "/admin/abandoned-baskets"),
    (re.compile(r"spec/filter group scoping|specifications group scoping", re.I), "/admin/specifications"),

    # ----- Order tracking & customer comms -----
    (re.compile(r"track order|public track|case-insensitive|auto-track|note from our team|did you mean", re.I), "/admin/online-orders"),
    (re.compile(r"order confirmation email|order status update email|custom email from admin|email history|last contacted", re.I), "/admin/online-orders"),

    # ----- Sample service -----
    (re.compile(r"tilesampleservicepage|sample service cms|sample service|/shop/tile-sample-service", re.I), "/admin/sample-service-content"),

    # ----- Customer accounts -----
    (re.compile(r"customer register|customer login|customer account|order history|forgot password|reset password", re.I), "/admin/customer-account-settings"),

    # ----- Trade accounts -----
    (re.compile(r"trade login|trade register|trade account dashboard|welcome back|trade discount % \+ tier|order again|activity stream|cart auto-reprice|per-order savings pill|trade tier ladder", re.I), "/admin/trade-account-settings"),
    (re.compile(r"trade accounts list|trade list|customer pricing|customer invites|loyalty", re.I), "/admin/trade-accounts"),

    # ----- Marketing & recovery -----
    (re.compile(r"refer-a-friend|friend-xxxxxx|/shop/refer", re.I), "/admin/abandoned-baskets"),
    (re.compile(r"abandoned cart|abandoned basket|abandoned-cart whatsapp", re.I), "/admin/abandoned-baskets"),
    (re.compile(r"promo code", re.I), "/admin/promo-codes"),
    (re.compile(r"weekly digest", re.I), "/admin/weekly-digest"),
    (re.compile(r"storefront messages|trade login.*toast|trade logout", re.I), "/admin/storefront-features"),
    (re.compile(r"marketing funnel tile|marketing funnel card", re.I), "/admin"),

    # ----- Static / info pages -----
    (re.compile(r"info page|privacy|terms|faq|returns?|delivery info|backwards-compat redirect", re.I), "/admin/info-pages"),
    (re.compile(r"contact page|/shop/contact", re.I), "/admin/website-settings"),
    (re.compile(r"showroom epos|cash counter|standard epos|epos terminal|barcode/sku", re.I), "/admin/showroom-epos"),
    (re.compile(r"showrooms?(?!.*manager)|stores", re.I), "/admin/showrooms-manager"),

    # ----- Admin landing pages (manage = themselves) -----
    (re.compile(r"website hub", re.I), "/admin/website-hub"),
    (re.compile(r"communication hub", re.I), "/admin/communication-hub"),
    (re.compile(r"customers hub", re.I), "/admin/customers-hub"),
    (re.compile(r"settings hub", re.I), "/admin/settings-hub"),
    (re.compile(r"reports hub", re.I), "/admin/reports-hub"),
    (re.compile(r"sales hub", re.I), "/admin/sales-hub"),
    (re.compile(r"stock hub", re.I), "/admin/stock-hub"),

    # ----- Admin pages — manage at = themselves -----
    (re.compile(r"online orders|today at a glance|order detail modal|status update modal|send custom email|method column|new-order audible", re.I), "/admin/online-orders"),
    (re.compile(r"navigation menu", re.I), "/admin/navigation-menu"),
    (re.compile(r"website categories", re.I), "/admin/website-categories"),
    (re.compile(r"website filters|specifications group", re.I), "/admin/filters"),
    (re.compile(r"info pages editor", re.I), "/admin/info-pages"),
    (re.compile(r"storefront features", re.I), "/admin/storefront-features"),
    (re.compile(r"welcome popup", re.I), "/admin/welcome-popup"),
    (re.compile(r"page maintenance", re.I), "/admin/page-maintenance"),
    (re.compile(r"maintenance tasks|db health snapshot|ui health|critical ui health|migrator|hyphen", re.I), "/admin/maintenance"),
    (re.compile(r"bathroom page admin", re.I), "/admin/bathroom-page"),
    (re.compile(r"sample service.*(admin|cms)", re.I), "/admin/sample-service-content"),
    (re.compile(r"collection.*(manager|mappings|page settings|detail settings)", re.I), "/admin/collections-hub"),
    (re.compile(r"abandoned baskets", re.I), "/admin/abandoned-baskets"),
    (re.compile(r"whatsapp", re.I), "/admin/whatsapp"),
    (re.compile(r"email inbox", re.I), "/admin/email-inbox"),
    (re.compile(r"email composer", re.I), "/admin/email-composer"),
    (re.compile(r"notification settings", re.I), "/admin/notifications"),
    (re.compile(r"live chat", re.I), "/admin/live-chat"),
    (re.compile(r"live visitors", re.I), "/admin/live-visitors"),
    (re.compile(r"website analytics", re.I), "/admin/website-analytics"),
    (re.compile(r"website sales dashboard", re.I), "/admin/website-sales"),
    (re.compile(r"permissions admin|/admin/permissions|roles\b|column-level rbac", re.I), "/admin/permissions"),

    # ----- EPOS -----
    (re.compile(r"showroom epos|cash counter|standard epos|trade pricing applied|trade tier discount|credit back accrual|volume pricing tier compatibility|multi-showroom support", re.I), "/admin/showroom-epos"),
    (re.compile(r"cash quotation", re.I), "/admin/cash-quotation"),
    (re.compile(r"daily.*reconciliation|email me this|reconciliation audit|reconciliation email|z-?read", re.I), "/admin/sales-hub"),

    # ----- Invoicing -----
    (re.compile(r"invoice history(?!.*showroom)", re.I), "/admin/invoice-history"),
    (re.compile(r"showroom invoice", re.I), "/admin/showroom-invoice-history"),
    (re.compile(r"invoice\b(?!.*history)", re.I), "/admin/invoice"),
    (re.compile(r"proforma", re.I), "/admin/proforma-invoice"),
    (re.compile(r"quotation history", re.I), "/admin/quotation-history"),
    (re.compile(r"quotation\b(?!.*history)", re.I), "/admin/quotation"),
    (re.compile(r"refund history", re.I), "/admin/refund-history"),
    (re.compile(r"refund\b(?!.*history)", re.I), "/admin/refund"),
    (re.compile(r"credit note history", re.I), "/admin/credit-note-history"),
    (re.compile(r"credit note\b(?!.*history)", re.I), "/admin/credit-note"),
    (re.compile(r"document storage|soft-deleted invoices", re.I), "/admin/document-storage"),

    # ----- Stock -----
    (re.compile(r"stock allocation", re.I), "/admin/stock-allocation"),
    (re.compile(r"stock transfers", re.I), "/admin/stock-transfers"),
    (re.compile(r"stock import", re.I), "/admin/stock-import"),
    (re.compile(r"bulk stock", re.I), "/admin/bulk-stock-edit"),
    (re.compile(r"stocktake", re.I), "/admin/stocktake-report"),
    (re.compile(r"stock cost", re.I), "/admin/stock-cost-report"),
    (re.compile(r"reorder suggestions", re.I), "/admin/reorder-suggestions"),
    (re.compile(r"to-?order report", re.I), "/admin/to-order-report"),
    (re.compile(r"batch tracking", re.I), "/admin/batch-tracking"),
    (re.compile(r"delivery check", re.I), "/admin/delivery-check-in"),
    (re.compile(r"delivery management", re.I), "/admin/delivery-management"),
    (re.compile(r"always-in-stock", re.I), "/admin/supplier-products"),

    # ----- CRM / customers admin -----
    (re.compile(r"quote requests", re.I), "/admin/quote-requests"),
    (re.compile(r"bulk inquiries", re.I), "/admin/bulk-inquiries"),
    (re.compile(r"tasks.*notes", re.I), "/admin/tasks-notes"),

    # ----- Catalogue / products admin -----
    (re.compile(r"supplier products|smart select|quick actions bar|bulk category|bulk pricing|sale & labels|tier pricing|product labels|name lock|description generator|product documents|quick edit|bulk edit history|manage options|manual create|product form", re.I), "/admin/supplier-products"),
    (re.compile(r"specifications manager", re.I), "/admin/specifications"),
    (re.compile(r"categor(y|ies) manager|manage categories", re.I), "/admin/categories"),
    (re.compile(r"clearance products", re.I), "/admin/clearance-products"),
    (re.compile(r"new collection products", re.I), "/admin/new-collection-products"),
    (re.compile(r"trash", re.I), "/admin/trash"),
    (re.compile(r"essentials needing photos", re.I), "/admin/sales-hub"),

    # ----- Sync & suppliers -----
    (re.compile(r"sync hub|supplier sync", re.I), "/admin/sync-hub"),
    (re.compile(r"supplier scrapers|force stop|validation warnings|null sku|cleanup orphaned", re.I), "/admin/supplier-sync-dashboard"),
    (re.compile(r"\bsuppliers? list\b|^suppliers", re.I), "/admin/suppliers"),
    (re.compile(r"supplier health", re.I), "/admin/supplier-health"),
    (re.compile(r"supplier images|image migration|image scraper|leporce|plus39", re.I), "/admin/supplier-images"),
    (re.compile(r"wallcano price", re.I), "/admin/wallcano-price-import"),
    (re.compile(r"scraping portal", re.I), "/admin/scraping-portal"),

    # ----- Reports / dashboards -----
    (re.compile(r"reports page|reports.*revenue|weekly compare|monthly compare", re.I), "/admin/reports"),
    (re.compile(r"historical sales", re.I), "/admin/historical-sales"),
    (re.compile(r"showroom dashboard", re.I), "/admin/showroom-dashboard"),
    (re.compile(r"staff performance", re.I), "/admin/staff-performance-dashboard"),
    (re.compile(r"audit trail", re.I), "/admin/audit-trail"),
    (re.compile(r"legacy orders|order management workflow", re.I), "/admin/orders"),

    # ----- Staff & auth -----
    (re.compile(r"admin auth", re.I), "/admin/auth"),
    (re.compile(r"staff register|staff invites", re.I), "/admin/staff-invites"),
    (re.compile(r"staff pins", re.I), "/admin/staff-pins"),
    (re.compile(r"user management", re.I), "/admin/user-management"),
    (re.compile(r"security settings", re.I), "/admin/security-settings"),
    (re.compile(r"staff chat", re.I), "/admin/staff-chat"),

    # ----- Showrooms & settings -----
    (re.compile(r"showrooms manager|showrooms? \(crud\)", re.I), "/admin/showrooms-manager"),
    (re.compile(r"pricing settings", re.I), "/admin/pricing-settings"),
    (re.compile(r"tiles info", re.I), "/admin/tiles-info"),
    (re.compile(r"trade account settings", re.I), "/admin/trade-account-settings"),
    (re.compile(r"checkout settings", re.I), "/admin/checkout-settings"),
    (re.compile(r"collection detail settings", re.I), "/admin/collection-detail-settings"),
    (re.compile(r"collections page settings", re.I), "/admin/collections-page-settings"),
    (re.compile(r"price tickets", re.I), "/admin/price-tickets"),
    (re.compile(r"general\b.*setting|^settings page", re.I), "/admin/settings"),

    # ----- Plumbing / migrators -----
    (re.compile(r"savings_meta|trade reprice toasts|auto-forward|data-testid|box metadata backfill|resend transactional", re.I), "/admin/maintenance"),
    (re.compile(r"co-purchase cache|nightly cron|abandoned cart reminders cron|weekly digest cron", re.I), "/admin/maintenance"),
    (re.compile(r"showroom id backfill|legacy /shop/tiles|tools \+ accessories merge|hyphen→space|hyphen->space", re.I), "/admin/maintenance"),
]


def resolve_admin_path(feature_text: str) -> str:
    for pattern, path in ADMIN_RULES:
        if pattern.search(feature_text):
            return path
    return "—"


# =============================================================================
# VERIFIED-BY-E1 set (PRD-derived; all features that were e2e tested or curl-
# verified during this session). Keyed by case-insensitive substring match on
# the feature text. Conservative — only marks rows the PRD explicitly verified.
# =============================================================================
VERIFIED_PATTERNS = [
    # Cart & checkout — verified by testing agent (iter 117/118/121/122/126/127) + Playwright
    "free-delivery progress",
    "box-sized stepper",
    "tier upsell nudge",
    "volume tier discount",
    "per-line tier savings",
    "cart-level total savings",
    "free sample upsell",
    "cart save banner",
    "promo code field",
    "3-step checkout",
    "billing address",
    "click & collect",
    "stripe hosted checkout",
    "stripe — klarna",
    "klarna express",
    "paypal express",
    "apple pay + google pay",
    "order success page",
    "tolerant guest-checkout schema",
    # Tracking
    "public track order",
    "case-insensitive lookup",
    "auto-track on url params",
    "note from our team",
    "did you mean",
    # PDP
    "dynamic proportional size pills",
    "volume pricing tier table",
    "trade customer box",
    "add-to-cart with box-step",
    "frequently-bought-together",
    "show actual size",
    "trade bundle booster",
    "free-sample cta inside",
    "sale ribbon + bottom badge",
    # Trade
    "trade login",
    "trade account dashboard",
    "welcome back",
    "trade discount % + tier badge",
    "order again",
    "activity stream",
    "cart auto-reprice",
    "per-order savings pill",
    "trade login (/shop/trade/login)",
    "trade customer box (login cta)",
    # Marketing
    "welcome popup",
    "refer-a-friend",
    "abandoned cart sequence",
    "abandoned cart whatsapp",
    "promo codes admin",
    "marketing funnel tile",
    "weekly digest email",
    "storefront messages admin",
    # Comms
    "order confirmation email",
    "order status update emails",
    "custom email from admin",
    "email history panel",
    # Static / homepage
    "editable homepage category tiles",
    "maintenance advance-notice ribbon",
    "whole-site maintenance switch",
    "tile collections page",
    "dynamic page h1",
    "backwards-compat redirects",
    # Admin website
    "homepage manager",
    "page maintenance admin",
    "maintenance tasks",
    "critical ui health panel",
    "site map manager",
    "tools+accessories db consolidation",
    # Online orders
    "online orders list",
    "today at a glance",
    "order detail modal",
    "status update modal",
    "send custom email button",
    "method column correctly",
    # RBAC + analytics
    "permissions admin",
    "live visitors page",
    "marketing funnel card",
    # Plumbing
    "free-delivery threshold consolidated",
    "savings_meta persisted",
    "trade reprice toasts",
    "auto-forward on trade",
    "data-testid coverage",
    "box metadata backfill",
    "resend transactional email",
    # Storefront pages added to UI health checks Apr 27
    "tile collections page",
    "static / info pages",
    "contact page",
    "showrooms / stores",
    "backwards-compat redirects",
    "sample basket (3-sample limit)",
    "sample service cms",
    "tile calculator page",
    "wishlist (",
    "compare products tray",
    "customer register",
    "customer login",
    "trade register",
    "15-minute checkout maintenance",
    # EPOS
    "daily sales reconciliation",
    "email me this",
    "auto-scheduled daily reconciliation",
    "reconciliation audit log",
    "soft-deleted invoices excluded",
    "smart select toolbar",
    "floating quick actions bar",
    "bulk pricing unit",
    "custom tier pricing per product",
    "sale & labels modal",
    "custom product labels",
    "spec/filter group scoping",
    "showroom id backfill",
    "site maintenance window cron",
    "abandoned cart reminders cron",
    "weekly digest cron",
    "co-purchase cache nightly",
    "hyphen→space url migrator",
    "tools + accessories merge migrator",
    "db health snapshot",
    "critical ui health checks",
]


def is_verified(feature_text: str) -> bool:
    lo = feature_text.lower()
    return any(p in lo for p in VERIFIED_PATTERNS)


# =============================================================================
# EXCEL BUILDER
# =============================================================================
def build_excel(filename: str, title: str, rows: list[tuple[str, str, str]], note: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="1E293B")
    section_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="475569")
    head_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="334155")
    body_font = Font(name="Calibri", size=10)
    note_font = Font(name="Calibri", size=9, italic=True, color="64748B")
    mono_font = Font(name="Consolas", size=9, color="0F172A")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    p_fills = {
        "P0": PatternFill("solid", fgColor="FEE2E2"),
        "P1": PatternFill("solid", fgColor="FEF3C7"),
        "P2": PatternFill("solid", fgColor="DBEAFE"),
    }
    verified_fill = PatternFill("solid", fgColor="DCFCE7")  # emerald-100
    verified_font = Font(name="Calibri", size=10, bold=True, color="166534")

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # Subtitle / note
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generated: {GENERATED_AT}   ·   {note}"
    ws["A2"].font = note_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    headers = ["#", "Section", "Feature", "Priority", "Path / URL", "Manage At (admin)", "Verified by E1", "User Sign-off", "Notes"]
    head_row = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=i, value=h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[head_row].height = 22

    cur_section = None
    r = head_row + 1
    idx = 0
    verified_count = 0
    for section, feature, priority in rows:
        if section != cur_section:
            cur_section = section
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
            sc = ws.cell(row=r, column=1, value=f"  {section}")
            sc.font = section_font
            sc.fill = section_fill
            sc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 20
            r += 1

        idx += 1
        path = resolve_path(feature)
        admin_path = resolve_admin_path(feature)
        verified = is_verified(feature)
        if verified:
            verified_count += 1

        ws.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2, value=section).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=r, column=3, value=feature).alignment = Alignment(vertical="center", wrap_text=True)

        pcell = ws.cell(row=r, column=4, value=priority)
        pcell.alignment = Alignment(horizontal="center", vertical="center")
        pcell.fill = p_fills.get(priority, PatternFill())
        pcell.font = Font(name="Calibri", size=10, bold=True)

        path_cell = ws.cell(row=r, column=5, value=path)
        path_cell.alignment = Alignment(vertical="center", wrap_text=True)
        path_cell.font = mono_font

        admin_cell = ws.cell(row=r, column=6, value=admin_path)
        admin_cell.alignment = Alignment(vertical="center", wrap_text=True)
        admin_cell.font = mono_font
        # Subtle indigo tint so the column visually pops as "where I edit this"
        admin_cell.fill = PatternFill("solid", fgColor="EEF2FF")

        vcell = ws.cell(row=r, column=7, value="✓" if verified else "")
        vcell.alignment = Alignment(horizontal="center", vertical="center")
        if verified:
            vcell.fill = verified_fill
            vcell.font = verified_font

        ws.cell(row=r, column=8, value="☐").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=9, value="").alignment = Alignment(vertical="center", wrap_text=True)

        for col in range(1, 10):
            ws.cell(row=r, column=col).border = border
            if col not in (4, 5, 6, 7):
                ws.cell(row=r, column=col).font = body_font
        ws.row_dimensions[r].height = 30
        r += 1

    # Column widths — # | Section | Feature | P | Path | Manage At | E1 | Sign-off | Notes
    widths = [5, 22, 52, 8, 30, 30, 12, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    # Legend sheet
    legend = wb.create_sheet("Legend")
    legend.column_dimensions["A"].width = 18
    legend.column_dimensions["B"].width = 90
    legend["A1"] = "Key"
    legend["B1"] = "Meaning"
    for c in (legend["A1"], legend["B1"]):
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center")
    legend_rows = [
        ("P0", "Launch-critical / revenue path / customer-facing must-have"),
        ("P1", "High value but non-blocking — fix soon after launch if anything slips"),
        ("P2", "Backlog / nice-to-have"),
        ("Path / URL", "Customer-facing path on production. Append to your Railway domain to open & verify."),
        ("Manage At", "Admin path where you edit / toggle / configure this feature. Always begins /admin/…"),
        ("Verified by E1 ✓", "Pre-ticked items were e2e tested by E1 (testing agent / Playwright / curl). Spot-check only — focus your time on un-ticked rows."),
        ("User Sign-off ☐", "Empty checkbox — tick once you've personally clicked through and confirmed live on production."),
        (f"Total verified", f"{verified_count} of {len(rows)} rows pre-ticked ({(verified_count/len(rows)*100):.0f}%)"),
    ]
    for i, (k, v) in enumerate(legend_rows, start=2):
        kcell = legend.cell(row=i, column=1, value=k)
        kcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        kcell.font = Font(bold=True)
        if k in p_fills:
            kcell.fill = p_fills[k]
        elif k.startswith("Verified"):
            kcell.fill = verified_fill
            kcell.font = verified_font
        elif k == "Manage At":
            kcell.fill = PatternFill("solid", fgColor="EEF2FF")
        legend.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="center")
        legend.row_dimensions[i].height = 26
    legend.sheet_view.showGridLines = False

    out = OUT_DIR / filename
    wb.save(out)
    return out, verified_count


# =============================================================================
# PDF BUILDER
# =============================================================================
def build_pdf(filename: str, title: str, rows: list[tuple[str, str, str]], note: str):
    out = OUT_DIR / filename
    # Landscape A4 (297×210 mm) — needed to fit the extra "Manage At" column
    # without cramping the Feature description.
    doc = SimpleDocTemplate(
        str(out),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleX", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=16, textColor=colors.HexColor("#0F172A"), spaceAfter=2, alignment=0,
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"),
        spaceAfter=8,
    )
    section_style = ParagraphStyle(
        "Sect", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=10, textColor=colors.white, leftIndent=4,
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontSize=7.8, leading=10,
        textColor=colors.HexColor("#0F172A"),
    )
    path_style = ParagraphStyle(
        "Path", parent=styles["Normal"], fontName="Courier",
        fontSize=7, leading=9, textColor=colors.HexColor("#0F172A"),
    )

    verified_count = sum(1 for _, f, _ in rows if is_verified(f))

    story = [
        Paragraph(title, title_style),
        Paragraph(
            f"Generated: {GENERATED_AT} &nbsp;·&nbsp; {note} &nbsp;·&nbsp; "
            f"<b>{verified_count}/{len(rows)}</b> pre-verified by E1",
            sub_style,
        ),
    ]

    p_colors = {
        "P0": colors.HexColor("#FECACA"),
        "P1": colors.HexColor("#FDE68A"),
        "P2": colors.HexColor("#BFDBFE"),
    }
    verified_bg = colors.HexColor("#DCFCE7")

    grouped: dict[str, list[tuple[str, str]]] = {}
    section_order: list[str] = []
    for section, feature, priority in rows:
        if section not in grouped:
            grouped[section] = []
            section_order.append(section)
        grouped[section].append((feature, priority))

    idx = 0
    page_w = landscape(A4)[0] - 20 * mm  # 277 mm usable width
    col_widths = [
        7 * mm,    # #
        85 * mm,   # Feature
        9 * mm,    # P
        56 * mm,   # Path
        56 * mm,   # Manage At
        12 * mm,   # E1 ✓
        12 * mm,   # ☐
        40 * mm,   # Notes
    ]
    assert sum(col_widths) <= page_w + 0.5

    admin_bg = colors.HexColor("#EEF2FF")  # subtle indigo tint for "Manage At" column

    for section in section_order:
        banner = RLTable(
            [[Paragraph(section, section_style)]],
            colWidths=[sum(col_widths)],
        )
        banner.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#475569")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(banner)

        data = [["#", "Feature", "P", "Path / URL", "Manage At (admin)", "E1 ✓", "Done", "Notes"]]
        styles_rules = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("ALIGN", (2, 1), (2, -1), "CENTER"),
            ("ALIGN", (5, 1), (6, -1), "CENTER"),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 1), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 2),
            # Persistent indigo tint on the entire Manage At column body
            ("BACKGROUND", (4, 1), (4, -1), admin_bg),
        ]
        for feature, priority in grouped[section]:
            idx += 1
            path = resolve_path(feature)
            admin_path = resolve_admin_path(feature)
            verified = is_verified(feature)
            data.append([
                str(idx),
                Paragraph(feature, cell_style),
                priority,
                Paragraph(path, path_style),
                Paragraph(admin_path, path_style),
                "✓" if verified else "",
                "☐",
                "",
            ])
            row_idx = len(data) - 1
            styles_rules.append((
                "BACKGROUND", (2, row_idx), (2, row_idx),
                p_colors.get(priority, colors.white),
            ))
            styles_rules.append(("FONTNAME", (2, row_idx), (2, row_idx), "Helvetica-Bold"))
            if verified:
                styles_rules.append(("BACKGROUND", (5, row_idx), (5, row_idx), verified_bg))
                styles_rules.append(("FONTNAME", (5, row_idx), (5, row_idx), "Helvetica-Bold"))
                styles_rules.append(("TEXTCOLOR", (5, row_idx), (5, row_idx), colors.HexColor("#166534")))

        tbl = RLTable(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle(styles_rules))
        story.append(tbl)
        story.append(Spacer(1, 4))

    # Legend
    story.append(Spacer(1, 5))
    legend_data = [
        ["Legend", ""],
        ["P0", "Launch-critical / revenue path / customer-facing must-have"],
        ["P1", "High value but non-blocking — fix soon after launch if anything slips"],
        ["P2", "Backlog / nice-to-have"],
        ["E1 ✓", "Pre-ticked: e2e verified by E1 (testing agent / Playwright / curl). Spot-check only."],
        ["Done ☐", "Tick once you've personally confirmed live on production."],
        ["Path", "Customer-facing path on production. Append to your Railway domain (e.g. https://www.tilestation.co.uk + path)."],
        ["Manage At", "Admin path where you edit / toggle / configure this feature. Always begins /admin/…"],
    ]
    legend_tbl = RLTable(legend_data, colWidths=[22 * mm, sum(col_widths) - 22 * mm])
    legend_tbl.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, 0), 6),
        ("BACKGROUND", (0, 1), (0, 1), p_colors["P0"]),
        ("BACKGROUND", (0, 2), (0, 2), p_colors["P1"]),
        ("BACKGROUND", (0, 3), (0, 3), p_colors["P2"]),
        ("BACKGROUND", (0, 4), (0, 4), verified_bg),
        ("BACKGROUND", (0, 7), (0, 7), admin_bg),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (1, 1), (1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(legend_tbl)

    doc.build(story)
    return out, verified_count


def main():
    deliverables = [
        ("Website_Features_Checklist", "Tile Station — Website / Storefront Features Checklist",
         WEBSITE_FEATURES,
         "Customer-facing storefront, marketing funnel, customer/trade accounts, admin website management"),
        ("EPOS_Features_Checklist", "Tile Station — EPOS / Back-Office Features Checklist",
         EPOS_FEATURES,
         "In-store EPOS, invoicing, stock, trade management, products catalogue, sync hub, reports"),
    ]

    print("\n=== Generated checklists ===")
    for stem, title, rows, note in deliverables:
        xlsx, vc_x = build_excel(f"{stem}.xlsx", title, rows, note)
        pdf, vc_p = build_pdf(f"{stem}.pdf", title, rows, note)
        for f in (xlsx, pdf):
            shutil.copy(f, PUBLIC_DIR / f.name)
            size_kb = f.stat().st_size / 1024
            print(f"  {f}  ({size_kb:.1f} KB)  →  {PUBLIC_DIR / f.name}")
        print(f"     ({title}) — {vc_x}/{len(rows)} pre-verified")


if __name__ == "__main__":
    main()
