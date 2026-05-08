"""
Stock Import API - Update showroom stock from spreadsheet data
"""
import json
from datetime import datetime, timezone
from typing import List, Optional
from fastapi import APIRouter, HTTPException, Depends, File, UploadFile
from pydantic import BaseModel
import openpyxl
import io

from config import get_db
from services import get_current_user, require_admin_access, log_audit

router = APIRouter(tags=["Stock Import"])

SHOWROOM_IDS = {
    "gravesend": "7411e9cf-2250-4bdc-a6ce-2dc844e32646",
    "tonbridge": "abd4e71c-ee0f-4bd1-9806-348cede6b696",
    "chingford": "57c16bc8-9538-46db-9d8b-9f7270d5c7b7",
    "sydenham": "6eab4459-aa87-4e87-aba4-4b3a44842879"
}


class StocktakeItem(BaseModel):
    supplier: str
    product: str
    size: Optional[str] = None
    stock: int
    price: Optional[float] = None
    code: Optional[str] = None
    cost: Optional[float] = None
    length: Optional[str] = None


class StocktakeUpload(BaseModel):
    showroom: str  # gravesend, tonbridge, chingford, sydenham
    items: List[StocktakeItem]
    dry_run: bool = True


def parse_stocktake_excel(file_content: bytes) -> List[dict]:
    """Parse stocktake Excel file and extract items"""
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    sheet = wb.active
    
    # Find header row
    header_row = None
    for row in range(1, min(10, sheet.max_row + 1)):
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row, col).value
            if val and str(val).strip().lower() == 'supplier':
                header_row = row
                break
        if header_row:
            break
    
    if not header_row:
        raise ValueError("Could not find header row with 'Supplier' column")
    
    # Get headers
    headers = {}
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(header_row, col).value
        if val:
            key = str(val).strip().lower()
            if 'stock' in key:
                key = 'stock'
            headers[key] = col
    
    # Extract data
    items = []
    for row in range(header_row + 1, sheet.max_row + 1):
        supplier = sheet.cell(row, headers.get('supplier', 1)).value
        product = sheet.cell(row, headers.get('product', 2)).value
        
        if not product or not supplier:
            continue
            
        size = sheet.cell(row, headers.get('size(s)', 3)).value if headers.get('size(s)') else None
        stock = sheet.cell(row, headers.get('stock', 4)).value
        price = sheet.cell(row, headers.get('list price', 5)).value if headers.get('list price') else None
        code = sheet.cell(row, headers.get('code', 0)).value if headers.get('code') else None
        cost = sheet.cell(row, headers.get('cost', 0)).value if headers.get('cost') else None
        length = sheet.cell(row, headers.get('length', 0)).value if headers.get('length') else None
        
        items.append({
            'supplier': str(supplier).strip(),
            'product': str(product).strip(),
            'size': str(size).strip() if size else None,
            'stock': int(stock) if stock is not None else 0,
            'price': float(price) if price else None,
            'code': str(code).strip() if code else None,
            'cost': float(cost) if cost else None,
            'length': str(length).strip() if length else None
        })
    
    wb.close()
    return items


def match_product(item: dict, db_products: list) -> tuple:
    """Try to match a stocktake item to a database product"""
    product_name = item['product'].lower()
    size = item.get('size', '').lower() if item.get('size') else ''
    code = item.get('code', '').upper() if item.get('code') else ''
    
    # Method 1: Match by code
    if code:
        for p in db_products:
            if p.get('sku', '').upper() == code:
                return p, 'sku'
    
    # Method 2: Match by name components
    search_words = [w.lower() for w in product_name.split() if len(w) > 2]
    
    best_match = None
    best_score = 0
    
    for p in db_products:
        db_name = (p.get('name') or '').lower()
        
        # Count matching words
        matches = sum(1 for w in search_words if w in db_name)
        score = matches / len(search_words) if search_words else 0
        
        # Size must match if specified
        if size:
            size_num = size.replace('mm', '').replace('ml', '').strip()
            if size_num not in db_name and size not in db_name:
                continue
            score += 0.3  # Bonus for size match
        
        if score > best_score:
            best_score = score
            best_match = p
    
    if best_match and best_score >= 0.6:
        return best_match, f'name_match_{int(best_score*100)}%'
    
    return None, None


@router.post("/stock-import/upload")
async def upload_stocktake_file(
    showroom: str,
    file: UploadFile = File(...),
    dry_run: bool = True,
    current_user: dict = Depends(get_current_user)
):
    """
    Upload a stocktake Excel file to update showroom stock levels.
    
    Args:
        showroom: One of: gravesend, tonbridge, chingford, sydenham
        file: Excel file (.xlsx) with columns: Supplier, Product, Size(s), Stock Available/Quantity, etc.
        dry_run: If True, only preview changes without applying them
    """
    require_admin_access(current_user)
    
    showroom_lower = showroom.lower()
    if showroom_lower not in SHOWROOM_IDS:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid showroom. Must be one of: {', '.join(SHOWROOM_IDS.keys())}"
        )
    
    showroom_id = SHOWROOM_IDS[showroom_lower]
    
    # Read and parse file
    try:
        content = await file.read()
        items = parse_stocktake_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    if not items:
        raise HTTPException(status_code=400, detail="No items found in file")
    
    db = get_db()
    
    # Get all products
    db_products = await db.products.find({}, {"_id": 0}).to_list(50000)
    
    # Match items to products
    updates = []
    not_found = []
    
    for item in items:
        matched_product, match_method = match_product(item, db_products)
        
        if matched_product:
            updates.append({
                'product_id': matched_product.get('id'),
                'product_name': matched_product.get('name'),
                'product_sku': matched_product.get('sku'),
                'stocktake_item': item,
                'match_method': match_method,
                'new_stock': item['stock']
            })
        else:
            not_found.append(item)
    
    # Apply updates if not dry run
    updated_count = 0
    if not dry_run:
        for update in updates:
            product_id = update['product_id']
            new_stock = update['new_stock']
            
            # Get current product
            product = await db.products.find_one({'id': product_id})
            if not product:
                continue
            
            # Update showroom stock
            showroom_stock = product.get('showroom_stock', [])
            if isinstance(showroom_stock, list):
                # Find or create showroom entry
                found = False
                for i, ss in enumerate(showroom_stock):
                    if ss.get('showroom_id') == showroom_id:
                        showroom_stock[i]['quantity'] = new_stock
                        found = True
                        break
                if not found:
                    showroom_stock.append({
                        'showroom_id': showroom_id,
                        'quantity': new_stock
                    })
            else:
                showroom_stock = [{'showroom_id': showroom_id, 'quantity': new_stock}]
            
            # Calculate total stock across showrooms
            total_stock = sum(ss.get('quantity', 0) for ss in showroom_stock)
            
            # Update product
            await db.products.update_one(
                {'id': product_id},
                {
                    '$set': {
                        'showroom_stock': showroom_stock,
                        'stock': total_stock,
                        'updated_at': datetime.now(timezone.utc)
                    }
                }
            )
            updated_count += 1
        
        # Log audit
        await log_audit(
            action="STOCK_IMPORT",
            entity_type="products",
            user=current_user,
            details=f"Updated {updated_count} products for {showroom.title()} from stocktake file"
        )
    
    return {
        "showroom": showroom.title(),
        "showroom_id": showroom_id,
        "file_name": file.filename,
        "total_items_in_file": len(items),
        "matched": len(updates),
        "not_found": len(not_found),
        "updated": updated_count if not dry_run else 0,
        "dry_run": dry_run,
        "updates_preview": [
            {
                "product_name": u['product_name'],
                "product_sku": u['product_sku'],
                "stocktake_name": u['stocktake_item']['product'],
                "stocktake_size": u['stocktake_item'].get('size'),
                "new_stock": u['new_stock'],
                "match_method": u['match_method']
            }
            for u in updates[:50]  # Limit preview
        ],
        "not_found_items": [
            {
                "supplier": nf['supplier'],
                "product": nf['product'],
                "size": nf.get('size'),
                "stock": nf['stock']
            }
            for nf in not_found[:30]  # Limit preview
        ]
    }


@router.post("/stock-import/json")
async def import_stock_from_json(
    data: StocktakeUpload,
    current_user: dict = Depends(get_current_user)
):
    """
    Import stock from JSON data (for programmatic updates).
    """
    require_admin_access(current_user)
    
    showroom_lower = data.showroom.lower()
    if showroom_lower not in SHOWROOM_IDS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid showroom. Must be one of: {', '.join(SHOWROOM_IDS.keys())}"
        )
    
    showroom_id = SHOWROOM_IDS[showroom_lower]
    
    db = get_db()
    db_products = await db.products.find({}, {"_id": 0}).to_list(50000)
    
    updates = []
    not_found = []
    
    for item in data.items:
        item_dict = item.dict()
        matched_product, match_method = match_product(item_dict, db_products)
        
        if matched_product:
            updates.append({
                'product_id': matched_product.get('id'),
                'product_name': matched_product.get('name'),
                'stocktake_item': item_dict,
                'match_method': match_method,
                'new_stock': item.stock
            })
        else:
            not_found.append(item_dict)
    
    updated_count = 0
    if not data.dry_run:
        for update in updates:
            product = await db.products.find_one({'id': update['product_id']})
            if not product:
                continue
            
            showroom_stock = product.get('showroom_stock', [])
            if isinstance(showroom_stock, list):
                found = False
                for i, ss in enumerate(showroom_stock):
                    if ss.get('showroom_id') == showroom_id:
                        showroom_stock[i]['quantity'] = update['new_stock']
                        found = True
                        break
                if not found:
                    showroom_stock.append({'showroom_id': showroom_id, 'quantity': update['new_stock']})
            else:
                showroom_stock = [{'showroom_id': showroom_id, 'quantity': update['new_stock']}]
            
            total_stock = sum(ss.get('quantity', 0) for ss in showroom_stock)
            
            await db.products.update_one(
                {'id': update['product_id']},
                {'$set': {'showroom_stock': showroom_stock, 'stock': total_stock, 'updated_at': datetime.now(timezone.utc)}}
            )
            updated_count += 1
        
        await log_audit(
            action="STOCK_IMPORT",
            entity_type="products",
            user=current_user,
            details=f"Updated {updated_count} products for {data.showroom.title()} via JSON import"
        )
    
    return {
        "showroom": data.showroom.title(),
        "matched": len(updates),
        "not_found": len(not_found),
        "updated": updated_count if not data.dry_run else 0,
        "dry_run": data.dry_run
    }
