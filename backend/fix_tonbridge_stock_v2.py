import openpyxl
import os
import requests
import json

PROD_URL = "https://tile-station-production.up.railway.app"
EMAIL = "qasim@tilestation.co.uk"
PASSWORD = os.environ.get("TILESTATION_ADMIN_PASSWORD", "")

TONBRIDGE_ID = "c16acbcc-13da-427a-8677-1dfce132c027"

def login():
    response = requests.post(f"{PROD_URL}/api/auth/login", 
        json={"email": EMAIL, "password": PASSWORD})
    return response.json().get("token")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def parse_spreadsheet(filepath, header_row=3):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    trims = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        product = sheet.cell(row=row, column=2).value
        code = sheet.cell(row=row, column=7).value
        if product and code:
            sku = f"TRI-{code}"
            trims[sku] = {
                'stock': int(sheet.cell(row=row, column=4).value or 0),
                'name': product,
            }
    return trims

def main():
    token = login()
    print("✓ Logged in")
    
    # Parse Tonbridge spreadsheet
    tonbridge_trims = parse_spreadsheet('/app/tonbridge_trims_new.xlsx')
    print(f"✓ Parsed {len(tonbridge_trims)} Tonbridge trims")
    
    # Get existing products
    response = requests.get(f"{PROD_URL}/api/products?limit=3000", headers=get_headers(token))
    products = {p['sku']: p for p in response.json() if p.get('sku')}
    
    # Prepare allocations list
    allocations = []
    for sku, trim_data in tonbridge_trims.items():
        if sku in products:
            p = products[sku]
            allocations.append({
                "showroom_id": TONBRIDGE_ID,
                "quantity": trim_data['stock']
            })
            
            # Use the dedicated showroom stock endpoint
            response = requests.put(
                f"{PROD_URL}/api/products/{p['id']}/showroom-stock",
                headers=get_headers(token),
                json={"allocations": [{"showroom_id": TONBRIDGE_ID, "quantity": trim_data['stock']}]}
            )
            
            if response.status_code == 200:
                print(f"  ✓ {p['name'][:40]} - Tonbridge: {trim_data['stock']}")
            else:
                print(f"  ✗ {p['name'][:40]} - Error: {response.text[:50]}")
    
    print(f"\n✓ Processed {len(tonbridge_trims)} trims")

if __name__ == "__main__":
    main()
