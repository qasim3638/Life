import openpyxl
import os
import requests
import json

PROD_URL = "https://tile-station-production.up.railway.app"
EMAIL = "qasim@tilestation.co.uk"
PASSWORD = os.environ.get("TILESTATION_ADMIN_PASSWORD", "")

CHINGFORD_ID = "6aa930df-a561-441e-949a-88514d85f2dc"
TONBRIDGE_ID = "c16acbcc-13da-427a-8677-1dfce132c027"

def login():
    response = requests.post(f"{PROD_URL}/api/auth/login", 
        json={"email": EMAIL, "password": PASSWORD})
    return response.json().get("token")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def parse_spreadsheet(filepath, header_row=3):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    trims = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        product = sheet.cell(row=row, column=2).value
        code = sheet.cell(row=row, column=7).value
        if product and code:
            sku = f"TRI-{code}"
            trims[sku] = {
                'stock': int(sheet.cell(row=row, column=4).value or 0),
                'name': product,
                'size': sheet.cell(row=row, column=3).value
            }
    return trims

def main():
    token = login()
    print("✓ Logged in")
    
    # Parse Tonbridge spreadsheet
    tonbridge_trims = parse_spreadsheet('/app/tonbridge_trims_new.xlsx')
    print(f"✓ Parsed {len(tonbridge_trims)} Tonbridge trims")
    
    # Get existing products
    response = requests.get(f"{PROD_URL}/api/products?limit=3000", headers=get_headers(token))
    products = {p['sku']: p for p in response.json() if p.get('sku')}
    
    updated = 0
    for sku, trim_data in tonbridge_trims.items():
        if sku in products:
            p = products[sku]
            showroom_stock = p.get('showroom_stock', {})
            
            # Update Tonbridge stock
            showroom_stock[TONBRIDGE_ID] = trim_data['stock']
            
            # Calculate total
            total_stock = sum(showroom_stock.values())
            
            update_data = {
                'name': p['name'],
                'sku': p['sku'],
                'description': p.get('description'),
                'price': p.get('price'),
                'cost': p.get('cost'),
                'stock': total_stock,
                'reorder_level': p.get('reorder_level', 5),
                'showroom_stock': showroom_stock,
            }
            
            response = requests.put(
                f"{PROD_URL}/api/products/{p['id']}",
                headers=get_headers(token),
                json=update_data
            )
            
            if response.status_code == 200:
                updated += 1
                print(f"  ✓ {p['name'][:50]} - Tonbridge: {trim_data['stock']}")
            else:
                print(f"  ✗ Failed: {p['name'][:50]}")
    
    print(f"\n=== Updated Tonbridge stock for {updated} trims ===")

if __name__ == "__main__":
    main()
