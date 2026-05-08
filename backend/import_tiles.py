import openpyxl
import os
import requests
import json
import re

PROD_URL = "https://tile-station-production.up.railway.app"
EMAIL = "qasim@tilestation.co.uk"
PASSWORD = os.environ.get("TILESTATION_ADMIN_PASSWORD", "")

CHINGFORD_ID = "6aa930df-a561-441e-949a-88514d85f2dc"

def login():
    response = requests.post(f"{PROD_URL}/api/auth/login", 
        json={"email": EMAIL, "password": PASSWORD})
    return response.json().get("token")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def parse_price(price_str):
    """Extract numeric price from string like '39.99m2' or '29.99 Each'"""
    if not price_str:
        return None
    price_str = str(price_str)
    # Remove non-numeric characters except decimal point
    match = re.search(r'(\d+\.?\d*)', price_str)
    if match:
        return float(match.group(1))
    return None

def create_product_name(tile):
    """Create full product name with finish and size"""
    name = tile['our_name']
    finish = tile['finish']
    size = tile['size']
    
    # Build name: "Product Name Finish Size"
    parts = [name]
    if finish and finish.lower() not in name.lower():
        parts.append(finish)
    if size:
        parts.append(size)
    
    return ' '.join(parts)

def create_sku(tile, index):
    """Create SKU for tile"""
    # Use supplier name initial + product name initial + size
    supplier = tile['supplier_name'][:3].upper() if tile['supplier_name'] else 'TIL'
    name_parts = tile['our_name'].split()
    name_init = ''.join([p[0].upper() for p in name_parts[:2]]) if name_parts else 'XX'
    size = tile['size'].replace('x', '')
    finish_init = tile['finish'][0].upper() if tile['finish'] else 'X'
    
    return f"TILE-{supplier}{name_init}{finish_init}-{size}"

def create_description(tile):
    """Create detailed description"""
    parts = []
    
    if tile['supplier_name']:
        parts.append(f"Supplier: {tile['supplier_name']}")
    if tile['size']:
        parts.append(f"Size: {tile['size']}")
    if tile['finish']:
        parts.append(f"Finish: {tile['finish']}")
    if tile['colour']:
        parts.append(f"Colour: {tile['colour']}")
    if tile['material']:
        parts.append(f"Material: {tile['material']}")
    if tile['thickness']:
        parts.append(f"Thickness: {tile['thickness']}")
    if tile['suitability']:
        parts.append(f"Suitability: {tile['suitability']}")
    if tile['rectified']:
        parts.append(f"Rectified: {tile['rectified']}")
    if tile['underfloor']:
        parts.append(f"Underfloor Heating: {tile['underfloor']}")
    
    parts.append(f"Category: {tile['category']}")
    
    return ' | '.join(parts)

def main():
    print("=== IMPORTING CHINGFORD TILES ===\n")
    
    token = login()
    print("✓ Logged in")
    
    # Parse spreadsheet
    wb = openpyxl.load_workbook('/app/chingford_tiles.xlsx', data_only=True)
    sheet = wb.active
    
    tiles = []
    header_row = 3
    
    for row in range(header_row + 1, sheet.max_row + 1):
        product_name = sheet.cell(row=row, column=2).value
        if product_name:
            tile = {
                'supplier_name': str(sheet.cell(row=row, column=1).value or '').strip(),
                'our_name': str(product_name).strip(),
                'finish': str(sheet.cell(row=row, column=3).value or '').strip(),
                'size': str(sheet.cell(row=row, column=4).value or '').strip(),
                'boxes': int(sheet.cell(row=row, column=5).value or 0),
                'stock_m2': float(sheet.cell(row=row, column=6).value or 0),
                'list_price': sheet.cell(row=row, column=7).value,
                'colour': str(sheet.cell(row=row, column=8).value or '').strip(),
                'rectified': str(sheet.cell(row=row, column=9).value or '').strip(),
                'underfloor': str(sheet.cell(row=row, column=10).value or '').strip(),
                'suitability': str(sheet.cell(row=row, column=11).value or '').strip(),
                'thickness': str(sheet.cell(row=row, column=12).value or '').strip(),
                'code': sheet.cell(row=row, column=13).value,
                'material': str(sheet.cell(row=row, column=14).value or '').strip(),
                'cost_m2': sheet.cell(row=row, column=15).value,
                'category': str(sheet.cell(row=row, column=16).value or 'Floor Tiles').strip(),
            }
            tiles.append(tile)
    
    print(f"✓ Parsed {len(tiles)} tiles from spreadsheet")
    
    # Get existing products
    response = requests.get(f"{PROD_URL}/api/products?limit=3000", headers=get_headers(token))
    existing_products = response.json()
    existing_by_name = {}
    for p in existing_products:
        # Create searchable key from name
        name_key = p['name'].lower().replace(' ', '').replace('-', '')
        existing_by_name[name_key] = p
    
    print(f"✓ Found {len(existing_products)} existing products")
    
    created = 0
    updated = 0
    errors = []
    
    for i, tile in enumerate(tiles, 1):
        product_name = create_product_name(tile)
        sku = create_sku(tile, i)
        
        # Check if product exists (by similar name)
        name_key = product_name.lower().replace(' ', '').replace('-', '')
        existing = existing_by_name.get(name_key)
        
        # Parse price
        price = parse_price(tile['list_price'])
        cost = parse_price(tile['cost_m2']) if tile['cost_m2'] else None
        
        # Calculate stock (using boxes as quantity for now)
        stock = tile['boxes']
        
        # Prepare showroom stock
        showroom_stock = {CHINGFORD_ID: stock} if stock > 0 else {}
        
        product_data = {
            'name': product_name,
            'sku': sku,
            'description': create_description(tile),
            'price': price if price else 0.0,
            'cost': cost if cost else 0.0,
            'stock': stock,
            'reorder_level': 5,
            'showroom_stock': showroom_stock,
            'box_m2_coverage': round(tile['stock_m2'] / tile['boxes'], 2) if tile['boxes'] > 0 else None,
        }
        
        try:
            if existing:
                # Update existing product
                existing_showroom_stock = existing.get('showroom_stock', {})
                existing_showroom_stock[CHINGFORD_ID] = stock
                product_data['showroom_stock'] = existing_showroom_stock
                product_data['stock'] = sum(existing_showroom_stock.values())
                
                response = requests.put(
                    f"{PROD_URL}/api/products/{existing['id']}",
                    headers=get_headers(token),
                    json=product_data
                )
                
                if response.status_code == 200:
                    updated += 1
                    print(f"  ✓ Updated: {product_name[:50]} | Stock: {stock} boxes")
                else:
                    errors.append(f"Update failed for {product_name}: {response.text[:100]}")
            else:
                # Create new product
                response = requests.post(
                    f"{PROD_URL}/api/products",
                    headers=get_headers(token),
                    json=product_data
                )
                
                if response.status_code in [200, 201]:
                    created += 1
                    print(f"  + Created: {product_name[:50]} | Stock: {stock} boxes | £{price}")
                else:
                    errors.append(f"Create failed for {product_name}: {response.text[:100]}")
                    
        except Exception as e:
            errors.append(f"Error processing {product_name}: {str(e)}")
    
    print(f"\n=== SUMMARY ===")
    print(f"Created: {created}")
    print(f"Updated: {updated}")
    print(f"Errors: {len(errors)}")
    
    if errors:
        print(f"\n=== ERRORS (first 10) ===")
        for e in errors[:10]:
            print(f"  - {e}")

if __name__ == "__main__":
    main()
