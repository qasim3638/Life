import openpyxl
import os
import requests
import json

PROD_URL = "https://tile-station-production.up.railway.app"
EMAIL = "qasim@tilestation.co.uk"
PASSWORD = os.environ.get("TILESTATION_ADMIN_PASSWORD", "")

CHINGFORD_ID = "6aa930df-a561-441e-949a-88514d85f2dc"
TONBRIDGE_ID = "c16acbcc-13da-427a-8677-1dfce132c027"

def login():
    response = requests.post(f"{PROD_URL}/api/auth/login", 
        json={"email": EMAIL, "password": PASSWORD})
    return response.json().get("token")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def parse_spreadsheet():
    wb = openpyxl.load_workbook('/app/tools_accessories.xlsx', data_only=True)
    sheet = wb.active
    
    products = []
    for row in range(4, sheet.max_row + 1):
        product = sheet.cell(row=row, column=2).value
        if product:
            code = sheet.cell(row=row, column=6).value
            products.append({
                'supplier': str(sheet.cell(row=row, column=1).value or '').strip(),
                'product': str(product).strip(),
                'chingford_stock': int(sheet.cell(row=row, column=3).value or 0),
                'tonbridge_stock': int(sheet.cell(row=row, column=4).value or 0),
                'cost': sheet.cell(row=row, column=5).value,
                'code': str(code).strip() if code else None,
                'list_price': sheet.cell(row=row, column=7).value,
                'category': sheet.cell(row=row, column=8).value or 'Tools & Accessories',
            })
    return products

def create_sku(code):
    """Create SKU from code - use TIL- prefix for Tile Rite products"""
    if code:
        return f"TIL-{code}"
    return None

def create_description(item):
    """Create description with supplier info"""
    parts = ["Supplier: Tile Rite"]
    parts.append(f"Category: {item['category']}")
    return " | ".join(parts)

def main():
    print("=== TOOLS & ACCESSORIES UPDATE ===\n")
    
    token = login()
    print("✓ Logged in")
    
    # Parse spreadsheet
    spreadsheet_products = parse_spreadsheet()
    print(f"✓ Parsed {len(spreadsheet_products)} products from spreadsheet")
    
    # Get existing products from database
    response = requests.get(f"{PROD_URL}/api/products?limit=3000", headers=get_headers(token))
    existing_products = response.json()
    existing_by_sku = {p['sku']: p for p in existing_products if p.get('sku')}
    print(f"✓ Found {len(existing_products)} existing products in database")
    
    created = 0
    updated = 0
    stock_updated = 0
    errors = []
    
    for item in spreadsheet_products:
        sku = create_sku(item['code'])
        if not sku:
            continue
            
        existing = existing_by_sku.get(sku)
        
        # Calculate total stock
        total_stock = item['chingford_stock'] + item['tonbridge_stock']
        
        # Prepare showroom stock
        showroom_stock = {}
        if item['chingford_stock'] > 0:
            showroom_stock[CHINGFORD_ID] = item['chingford_stock']
        if item['tonbridge_stock'] > 0:
            showroom_stock[TONBRIDGE_ID] = item['tonbridge_stock']
        
        try:
            if existing:
                # Update existing product - keep existing cost if new cost is None
                new_cost = item['cost'] if item['cost'] is not None else existing.get('cost')
                new_price = item['list_price'] if item['list_price'] is not None else existing.get('price')
                
                # Preserve existing showroom stock for other showrooms
                existing_showroom_stock = existing.get('showroom_stock', {})
                existing_showroom_stock[CHINGFORD_ID] = item['chingford_stock']
                existing_showroom_stock[TONBRIDGE_ID] = item['tonbridge_stock']
                
                new_total = sum(existing_showroom_stock.values())
                
                # Update product details
                update_data = {
                    'name': item['product'],
                    'sku': sku,
                    'description': create_description(item),
                    'price': float(new_price) if new_price else existing.get('price'),
                    'cost': float(new_cost) if new_cost else existing.get('cost'),
                    'stock': new_total,
                    'reorder_level': existing.get('reorder_level', 5),
                    'showroom_stock': existing_showroom_stock,
                }
                
                response = requests.put(
                    f"{PROD_URL}/api/products/{existing['id']}",
                    headers=get_headers(token),
                    json=update_data
                )
                
                if response.status_code == 200:
                    updated += 1
                    print(f"  ✓ Updated: {item['product'][:45]} | C:{item['chingford_stock']} T:{item['tonbridge_stock']}")
                else:
                    errors.append(f"Update failed for {sku}: {response.text[:100]}")
            else:
                # Create new product
                product_data = {
                    'name': item['product'],
                    'sku': sku,
                    'description': create_description(item),
                    'price': float(item['list_price']) if item['list_price'] else 0.0,
                    'cost': float(item['cost']) if item['cost'] else 0.0,
                    'stock': total_stock,
                    'reorder_level': 5,
                    'showroom_stock': showroom_stock,
                }
                
                response = requests.post(
                    f"{PROD_URL}/api/products",
                    headers=get_headers(token),
                    json=product_data
                )
                
                if response.status_code in [200, 201]:
                    created += 1
                    print(f"  + Created: {item['product'][:45]} | C:{item['chingford_stock']} T:{item['tonbridge_stock']}")
                else:
                    errors.append(f"Create failed for {sku}: {response.text[:100]}")
                    
        except Exception as e:
            errors.append(f"Error processing {sku}: {str(e)}")
    
    print(f"\n=== SUMMARY ===")
    print(f"Created: {created}")
    print(f"Updated: {updated}")
    print(f"Errors: {len(errors)}")
    
    if errors:
        print(f"\n=== ERRORS (first 10) ===")
        for e in errors[:10]:
            print(f"  - {e}")

if __name__ == "__main__":
    main()
