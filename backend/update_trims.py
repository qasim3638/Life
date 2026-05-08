import openpyxl
import os
import requests
import json

PROD_URL = "https://tile-station-production.up.railway.app"
EMAIL = "qasim@tilestation.co.uk"
PASSWORD = os.environ.get("TILESTATION_ADMIN_PASSWORD", "")

def login():
    """Login and get token"""
    response = requests.post(f"{PROD_URL}/api/auth/login", 
        json={"email": EMAIL, "password": PASSWORD})
    if response.status_code == 200:
        return response.json().get("token")
    raise Exception(f"Login failed: {response.text}")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def get_showrooms(token):
    """Get showroom IDs"""
    response = requests.get(f"{PROD_URL}/api/showrooms", headers=get_headers(token))
    showrooms = {}
    for s in response.json():
        showrooms[s['name'].lower()] = s['id']
    return showrooms

def get_all_products(token):
    """Get all existing products"""
    response = requests.get(f"{PROD_URL}/api/products?limit=3000", headers=get_headers(token))
    return response.json()

def parse_spreadsheet(filepath, header_row=3):
    """Parse trim spreadsheet"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    trims = []
    for row in range(header_row + 1, sheet.max_row + 1):
        product = sheet.cell(row=row, column=2).value
        if product:
            code = sheet.cell(row=row, column=7).value
            trim = {
                'supplier': sheet.cell(row=row, column=1).value or 'Trimline',
                'product': product,
                'size': str(sheet.cell(row=row, column=3).value or ''),
                'stock': int(sheet.cell(row=row, column=4).value or 0),
                'list_price': sheet.cell(row=row, column=5).value,
                'length': sheet.cell(row=row, column=6).value,
                'code': str(code) if code else None,
                'cost': sheet.cell(row=row, column=8).value,
                'category': sheet.cell(row=row, column=9).value or 'Trims',
            }
            trims.append(trim)
    return trims

def create_product_name(product_name, size):
    """Create full product name with size"""
    size_str = size if size else ''
    # Check if size is already in the product name
    if size_str and size_str.lower() not in product_name.lower():
        return f"{product_name} {size_str}"
    return product_name

def create_sku(code):
    """Create SKU from code"""
    if code:
        return f"TRI-{code}"
    return None

def create_description(trim):
    """Create description with all relevant info"""
    parts = [f"Supplier: Trimline"]
    if trim['size']:
        parts.append(f"Size: {trim['size']}")
    if trim['length']:
        parts.append(f"Length: {trim['length']}")
    parts.append("Category: Trims")
    return " | ".join(parts)

def main():
    print("=== TRIM UPDATE SCRIPT ===\n")
    
    # Login
    token = login()
    print("✓ Logged in successfully")
    
    # Get showroom IDs
    showrooms = get_showrooms(token)
    chingford_id = showrooms.get('chingford')
    tonbridge_id = showrooms.get('tonbridge')
    print(f"✓ Showrooms: Chingford={chingford_id[:8]}..., Tonbridge={tonbridge_id[:8]}...")
    
    # Parse spreadsheets
    chingford_trims = parse_spreadsheet('/app/chingford_trims_new.xlsx')
    tonbridge_trims = parse_spreadsheet('/app/tonbridge_trims_new.xlsx')
    print(f"✓ Parsed {len(chingford_trims)} Chingford trims, {len(tonbridge_trims)} Tonbridge trims")
    
    # Get existing products
    existing_products = get_all_products(token)
    existing_by_sku = {p['sku']: p for p in existing_products if p.get('sku')}
    print(f"✓ Found {len(existing_products)} existing products")
    
    # Merge trim data by code (same product can be in both showrooms)
    all_trims = {}
    
    # Add Chingford trims
    for t in chingford_trims:
        code = t['code']
        if code:
            if code not in all_trims:
                all_trims[code] = {
                    'data': t,
                    'chingford_stock': t['stock'],
                    'tonbridge_stock': 0
                }
            else:
                all_trims[code]['chingford_stock'] = t['stock']
    
    # Add Tonbridge trims
    for t in tonbridge_trims:
        code = t['code']
        if code:
            if code not in all_trims:
                all_trims[code] = {
                    'data': t,
                    'chingford_stock': 0,
                    'tonbridge_stock': t['stock']
                }
            else:
                all_trims[code]['tonbridge_stock'] = t['stock']
                # Update data if tonbridge has more info
                if t['cost'] and not all_trims[code]['data']['cost']:
                    all_trims[code]['data']['cost'] = t['cost']
                if t['list_price'] and not all_trims[code]['data']['list_price']:
                    all_trims[code]['data']['list_price'] = t['list_price']
    
    print(f"✓ Total unique trims to process: {len(all_trims)}")
    
    # Process each trim
    created = 0
    updated = 0
    errors = []
    
    for code, trim_info in all_trims.items():
        t = trim_info['data']
        sku = create_sku(code)
        product_name = create_product_name(t['product'], t['size'])
        
        # Calculate total stock
        total_stock = trim_info['chingford_stock'] + trim_info['tonbridge_stock']
        
        # Check if product exists
        existing = existing_by_sku.get(sku)
        
        # Prepare showroom stock
        showroom_stock = {}
        if chingford_id:
            showroom_stock[chingford_id] = trim_info['chingford_stock']
        if tonbridge_id:
            showroom_stock[tonbridge_id] = trim_info['tonbridge_stock']
        
        # Prepare product data
        product_data = {
            'name': product_name,
            'sku': sku,
            'description': create_description(t),
            'price': float(t['list_price']) if t['list_price'] else 0.0,
            'cost': float(t['cost']) if t['cost'] else 0.0,
            'stock': total_stock,
            'reorder_level': 5,
            'showroom_stock': showroom_stock,
        }
        
        try:
            if existing:
                # Update existing product - preserve other showroom stock
                existing_showroom_stock = existing.get('showroom_stock', {})
                # Only update Chingford and Tonbridge, preserve others
                if chingford_id:
                    existing_showroom_stock[chingford_id] = trim_info['chingford_stock']
                if tonbridge_id:
                    existing_showroom_stock[tonbridge_id] = trim_info['tonbridge_stock']
                product_data['showroom_stock'] = existing_showroom_stock
                product_data['stock'] = sum(existing_showroom_stock.values())
                
                response = requests.put(
                    f"{PROD_URL}/api/products/{existing['id']}",
                    headers=get_headers(token),
                    json=product_data
                )
                if response.status_code == 200:
                    updated += 1
                    print(f"  ✓ Updated: {product_name} (Stock: {product_data['stock']})")
                else:
                    errors.append(f"Update failed for {sku}: {response.text}")
            else:
                # Create new product
                response = requests.post(
                    f"{PROD_URL}/api/products",
                    headers=get_headers(token),
                    json=product_data
                )
                if response.status_code in [200, 201]:
                    created += 1
                    print(f"  + Created: {product_name} (Stock: {total_stock})")
                else:
                    errors.append(f"Create failed for {sku}: {response.text}")
        except Exception as e:
            errors.append(f"Error processing {sku}: {str(e)}")
    
    print(f"\n=== SUMMARY ===")
    print(f"Created: {created}")
    print(f"Updated: {updated}")
    print(f"Errors: {len(errors)}")
    
    if errors:
        print(f"\n=== ERRORS ===")
        for e in errors[:10]:
            print(f"  - {e}")

if __name__ == "__main__":
    main()
