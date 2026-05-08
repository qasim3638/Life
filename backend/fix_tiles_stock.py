import openpyxl
import os
import requests

PROD_URL = "https://tile-station-production.up.railway.app"
EMAIL = "qasim@tilestation.co.uk"
PASSWORD = os.environ.get("TILESTATION_ADMIN_PASSWORD", "")

CHINGFORD_ID = "6aa930df-a561-441e-949a-88514d85f2dc"

def login():
    response = requests.post(f"{PROD_URL}/api/auth/login", 
        json={"email": EMAIL, "password": PASSWORD})
    return response.json().get("token")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def main():
    token = login()
    print("✓ Logged in")
    
    # Parse spreadsheet to get stock by product name
    wb = openpyxl.load_workbook('/app/chingford_tiles.xlsx', data_only=True)
    sheet = wb.active
    
    stock_data = {}
    for row in range(4, sheet.max_row + 1):
        product_name = sheet.cell(row=row, column=2).value
        if product_name:
            # Build key similar to how we named products
            name = str(product_name).strip()
            finish = str(sheet.cell(row=row, column=3).value or '').strip()
            size = str(sheet.cell(row=row, column=4).value or '').strip()
            
            # Create full name
            full_name = name
            if finish and finish.lower() not in name.lower():
                full_name += f" {finish}"
            if size:
                full_name += f" {size}"
            
            boxes = int(sheet.cell(row=row, column=5).value or 0)
            stock_data[full_name.lower()] = boxes
    
    print(f"✓ Parsed {len(stock_data)} tiles from spreadsheet")
    
    # Get tile products from database
    response = requests.get(f"{PROD_URL}/api/products?limit=3000", headers=get_headers(token))
    products = response.json()
    
    tiles = [p for p in products if p.get('sku', '').startswith('TILE-')]
    print(f"✓ Found {len(tiles)} tiles in database")
    
    updated = 0
    for tile in tiles:
        name_key = tile['name'].lower()
        
        if name_key in stock_data:
            stock = stock_data[name_key]
            
            # Update using showroom stock endpoint
            response = requests.put(
                f"{PROD_URL}/api/products/{tile['id']}/showroom-stock",
                headers=get_headers(token),
                json={"allocations": [{"showroom_id": CHINGFORD_ID, "quantity": stock}]}
            )
            
            if response.status_code == 200:
                updated += 1
                if stock > 0:
                    print(f"  ✓ {tile['name'][:45]} - {stock} boxes")
    
    print(f"\n✓ Updated stock for {updated} tiles")

if __name__ == "__main__":
    main()
