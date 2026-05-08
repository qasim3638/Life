import openpyxl
import os
import requests

PROD_URL = "https://tile-station-production.up.railway.app"
EMAIL = "qasim@tilestation.co.uk"
PASSWORD = os.environ.get("TILESTATION_ADMIN_PASSWORD", "")

CHINGFORD_ID = "6aa930df-a561-441e-949a-88514d85f2dc"
TONBRIDGE_ID = "c16acbcc-13da-427a-8677-1dfce132c027"

def login():
    response = requests.post(f"{PROD_URL}/api/auth/login", 
        json={"email": EMAIL, "password": PASSWORD})
    return response.json().get("token")

def get_headers(token):
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

def main():
    token = login()
    print("✓ Logged in")
    
    # Parse spreadsheet
    wb = openpyxl.load_workbook('/app/tools_accessories.xlsx', data_only=True)
    sheet = wb.active
    
    spreadsheet_data = {}
    for row in range(4, sheet.max_row + 1):
        code = sheet.cell(row=row, column=6).value
        if code:
            sku = f"TIL-{str(code).strip()}"
            spreadsheet_data[sku] = {
                'name': sheet.cell(row=row, column=2).value,
                'chingford': int(sheet.cell(row=row, column=3).value or 0),
                'tonbridge': int(sheet.cell(row=row, column=4).value or 0),
            }
    
    print(f"✓ Parsed {len(spreadsheet_data)} products from spreadsheet")
    
    # Get existing products
    response = requests.get(f"{PROD_URL}/api/products?limit=3000", headers=get_headers(token))
    products = {p['sku']: p for p in response.json() if p.get('sku')}
    
    updated = 0
    for sku, data in spreadsheet_data.items():
        if sku in products:
            p = products[sku]
            
            # Build allocations for both showrooms
            allocations = []
            if data['chingford'] > 0 or data['tonbridge'] > 0:
                allocations.append({"showroom_id": CHINGFORD_ID, "quantity": data['chingford']})
                allocations.append({"showroom_id": TONBRIDGE_ID, "quantity": data['tonbridge']})
            
            if allocations:
                response = requests.put(
                    f"{PROD_URL}/api/products/{p['id']}/showroom-stock",
                    headers=get_headers(token),
                    json={"allocations": allocations}
                )
                
                if response.status_code == 200:
                    updated += 1
                    if data['tonbridge'] > 0:
                        print(f"  ✓ {data['name'][:40]} - C:{data['chingford']} T:{data['tonbridge']}")
    
    print(f"\n✓ Updated showroom stock for {updated} products")

if __name__ == "__main__":
    main()
